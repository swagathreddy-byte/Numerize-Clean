import pandas as pd
import numpy as np
from datetime import datetime,date
from dateutil.relativedelta import *
from time import strptime
import glob
import shutil
import openpyxl
import json
import collections
import slugify
import re
from functools import reduce
from zipfile import ZipFile
import collections
from pytz import timezone
#%%
import os
import simplejson

# Class to generate reports from excel file

class RepGen:
    def __init__(self,file_name):
        self.Convert_to_Frame(file_name)
        self.Is_Data_good()
        self.error=[]

    def Convert_to_Frame(self,file_name):
        #self.dfs = pd.read_excel(file_name,header=None)
        pass

    def Is_Data_good(self):
        pass
        #print (self.dfs.isnull().values.any())
        #print (self.dfs.isnull().sum())
    def pass_error(self,err):
        self.error.append(str(err))



# Subclass to generate a report of particluar format from excel file
# Here we evaluate specific outputs needed to generate this report.

##-------------------- REPORT 1 --Profit Statement---------------------------##
class Prof_Stat(RepGen):

    def __init__(self,file_name, file_name1,file_suffix=None):
        super().__init__(file_name)
        try:
            self.dfs = pd.read_excel(file_name,header=None)
            self.df1 = pd.read_excel(file_name,skiprows=4,usecols="A:D")
            self.client = pd.read_excel(file_name1)
            self.head=pd.read_excel(file_name,nrows=4,header=None)
            self.Do_Accounting(file_suffix)
            self.Set_Writer()
            self.Is_Data_good()
            self.Set_Layout()
            self.Save_to_excel()
            self.json_writer()
            self.status = 'Success'
            #print('------------------Execution is complete---------------------')
        except:
            # print(e)
            self.status = 'Failure'
            raise


    def Set_Writer(self):

        self.df1.to_excel(self.writer, index=False, sheet_name='Profit and Loss by Month',startrow=4)


        self.workbook = self.writer.book
        self.worksheet = self.writer.sheets['Profit and Loss by Month']

        self.page2 = self.writer.sheets['Page 2']
        self.page3 = self.writer.sheets['Page 3']

    def Do_Accounting(self,file_suffix):

        self.df1['Total'] = self.df1[self.df1.columns[1]] + self.df1[self.df1.columns[2]]
        self.df1.rename(columns={self.df1.columns[0] :'Particulars'}, inplace=True)
        self.df1['Particulars'] = self.df1['Particulars'].str.strip()
        #self.df1.rename(columns={self.df1.columns[1] :self.df1.columns[1].strftime('%B')}, inplace=True)
        #self.df1.rename(columns={self.df1.columns[2] :self.df1.columns[2].strftime('%B')}, inplace=True)

        #Page 3 Manipulations
        self.fname='media/uploads/'+file_suffix+'Profit_Statement.xlsx'
        self.writer = pd.ExcelWriter(self.fname, engine='xlsxwriter')
        mon1,mon2 = self.head.iloc[2,0].split(',')[0].split('-')
        mon1 = str(mon1) + '-Amount (Rs)'
        mon2 = str(mon2) + '-Amount (Rs)'

        sum1=[]
        sum2=[]
        self.table ={}
        count = 0
        #lst = self.client['P&L_Bucket'].unique()
        lst = ['Sales','Other income','Food cost','Conversion cost','Selling cost','Administration cost','Marketing cost','Finance cost','Personnel cost']
        skiprowval = 4
        self.no_tables = len(lst)
        for items in lst:
            self.title = pd.DataFrame()
            self.title = self.title.append([items])
            self.title.to_excel(self.writer, index=False, sheet_name='Page 3',startrow=skiprowval-1,header=None)

            self.box = pd.DataFrame(columns =['Particulars', mon1, mon2] )
            buckets = self.client[self.client['P&L_Bucket']==items]['Particulars'].to_list()
            bucket= [item.strip() for item in buckets]
            #print(bucket)
            s1sum=0
            s2sum=0
            count = count+1

            for things in bucket:

                try:
                    s1 = self.df1[self.df1['Particulars']==things].iloc[0,1]
                    s2 = self.df1[self.df1['Particulars']==things].iloc[0,2]
                    self.box =self.box.append({'Particulars':things, mon1:s1, mon2:s2},ignore_index=True)
                    if  not np.isnan(s1):
                        if 'Closing stock' in things:
                            s1=-s1
                        s1sum = s1sum+(s1)

                    if not np.isnan(s2):
                        if 'Closing stock' in things:
                            s2=-s2
                        s2sum = s2sum+(s2)

                except Exception as e:
                    self.pass_error(e)


            sum1.append(s1sum)
            sum2.append(s2sum)
            if items=='Sales':
                tot_sales1=s1sum
                tot_sales2=s2sum
            if items=='Food cost':
                tot_fc1=s1sum
                tot_fc2=s2sum
            self.box =self.box.append({'Particulars':'Total', mon1:s1sum, mon2:s2sum},ignore_index=True)
            if items=='Food cost':
                self.box =self.box.append({'Particulars':'Food %', mon1:100*tot_fc1/tot_sales1, mon2:100*tot_fc2/tot_sales2},ignore_index=True)
            self.box.to_excel(self.writer, index=False, sheet_name='Page 3',startrow=skiprowval)
            skiprowval = skiprowval+self.box.shape[0]+4
            #Json table export---------------
            table = 'table'
            temp={}
            if not self.table:
                self.table[table]=[]
            temp['name'] =items
            temp['no_cols'] ='3'
            temp['col-headers'] ={"Col1":"Particulars","Col2":mon1,"Col3":mon2}
            dummy = self.box
            dummy.columns = ['Col1', 'Col2', 'Col3']
            temp['records'] = dummy.to_dict(orient='records')
            self.table[table].append(temp)


        #Page 2 Manipulations
        #lst = lst.tolist()
        loc =lst.index('Other income')
        lst.insert(loc+1,'Total Income')
        sum1.insert(loc+1,sum(sum1[:loc+1]))
        sum2.insert(loc+1,sum(sum2[:loc+1]))
        lst.append('Total Expenses')
        sum1.append(sum(sum1[loc+2:]))
        sum2.append(sum(sum2[loc+2:]))
        p1 =  [x*100 / sum(sum1[:loc+1]) for x in sum1]
        p2 =  [x*100 / sum(sum2[:loc+1]) for x in sum2]
        lst.append('Net profit/loss')
        sum1.append(sum(sum1[:loc+1])-(sum1[-1]))
        sum2.append(sum(sum2[:loc+1])-(sum2[-1]))
        p1.append(100*sum1[-1]/sum(sum1[:loc+1]))
        p2.append(100*sum2[-1]/sum(sum2[:loc+1]))


        self.pg2 = pd.DataFrame({'Particulars':lst,'Amount(Rs)':sum1,'%of total income':p1,'Amount (Rs)':sum2,'% of total income':p2})

        line = pd.DataFrame({'Particulars':'','Amount(Rs)':'','%of total income':'','Amount (Rs)':'','% of total income':''}, index=[loc+2])
        self.pg2 = pd.concat([self.pg2.iloc[:loc+1], line, self.pg2.iloc[loc+1:]]).reset_index(drop=True)
        self.pg2 = pd.concat([self.pg2.iloc[:loc+3], line, self.pg2.iloc[loc+3:]]).reset_index(drop=True)
        loc = self.pg2.shape[0]
        self.pg2 = pd.concat([self.pg2.iloc[:loc-1], line, self.pg2.iloc[loc-1:]]).reset_index(drop=True)

        self.pg2.to_excel(self.writer, index=False, sheet_name='Page 2',startrow=4)
        mon1,mon2 = self.head.iloc[2,0].split(',')[0].split('-')
        self.heading = pd.DataFrame(columns=['',mon1,'',mon2])
        self.heading.to_excel(self.writer, index=False, sheet_name='Page 2',startrow=3)

    def Set_Layout(self):
        hide_gridlines(self.worksheet,2)
        cell = 'D'+str(self.pg2.shape[0]+4)
        insert_image(self.page2,'fc_logo.png',cell)

        fmt= self.workbook.add_format({'num_format': '#,##0.00 _€'})
        self.worksheet.set_column('B:D', 20,fmt)
        self.worksheet.set_column('A:A', 50,fmt)
        fm2= self.workbook.add_format({'num_format': '#,##0.0 _€'})
        self.page3.set_column('A:D', 30,fmt)
        self.page2.set_column('A:E',20)
        self.page2.set_column('C:C',20,fm2)
        self.page2.set_column('E:E',20,fm2)

        merge_cells(self.workbook,self.worksheet,'A1','D1', str(self.dfs.iloc[0,0]))
        merge_cells(self.workbook,self.worksheet,'A2','D2', str(self.dfs.iloc[1,0]))
        merge_cells(self.workbook,self.worksheet,'A3','D3', str(self.dfs.iloc[2,0]))

        merge_cells(self.workbook,self.page3,'A1','C1', str(self.dfs.iloc[0,0]))
        merge_cells(self.workbook,self.page3,'A2','C2', 'Breakup of Profit statement')
        merge_cells(self.workbook,self.page3,'A3','C3', str(self.dfs.iloc[2,0]))

        merge_cells(self.workbook,self.page2,'A1','E1', str(self.dfs.iloc[0,0]))
        merge_cells(self.workbook,self.page2,'A2','E2', 'Profit statement' )
        merge_cells(self.workbook,self.page2,'A3','E3', str(self.dfs.iloc[2,0]))

        fm= self.workbook.add_format({'border': 1,'align': 'center','bold':1})

        self.page2.merge_range('B4:C4', str(self.heading.columns[1]), fm)
        self.page2.merge_range('D4:E4', str(self.heading.columns[3]), fm)

        lst = self.df1.index[self.df1['Particulars']=='Total Income'][0]

        currency_format = self.workbook.add_format({'num_format': '"₹"* #,##0.00 _€','bold': True})
        self.worksheet.write(lst+5,3, self.df1.iloc[lst,3], currency_format)
        self.worksheet.write(lst+5,2, self.df1.iloc[lst,2], currency_format)
        self.worksheet.write(lst+5,1, self.df1.iloc[lst,1], currency_format)
        self.worksheet.write(lst+5,0, self.df1.iloc[lst,0], currency_format)

    def Save_to_excel(self):
        self.writer.save()

    def json_writer(self):
        head = self.head.dropna(axis= 'columns', how='all')
        self.result = {}

        #----------Header file----------------------------

        head_json = {
            "customer": head.iloc[0,0],
            "report_type": head.iloc[1,0],
            "month_year":  head.iloc[2,0],
            "month1":  head.iloc[2,0].split('-')[0].strip(),
            "month2":  head.iloc[2,0].split('-')[1].split(',')[0].strip(),
            "no_tables":str(self.no_tables),
        }
        self.result['Page3'] ={}
        self.result['Page3']['header']=head_json

        self.result['Page3'].update(self.table)

        #--------Profit loss statement -------------------
        head_json = {
            "customer": head.iloc[0,0],
            "report_type": head.iloc[1,0],
            "month_year":  head.iloc[2,0],
            "month1":  head.iloc[2,0].split('-')[0].strip(),
            "month2":  head.iloc[2,0].split('-')[1].split(',')[0].strip(),
            "no_tables":"1",
        }
        self.result['Profit and Loss by Month'] ={}
        self.result['Profit and Loss by Month']['header']=head_json
        dummy = self.df1
        dummy.columns = ['Col1', 'Col2', 'Col3','Col4']
        PL_json = dummy.to_dict(orient='records')

        self.result['Profit and Loss by Month']['table1']={}
        self.result['Profit and Loss by Month']['table1']['name'] =''
        self.result['Profit and Loss by Month']['table1']['no_cols'] ='4'
        self.result['Profit and Loss by Month']['table1']['col-headers'] ={"Col1":"particulars","Col2":"July Amount (Rs)","Col3":"August Amount(Rs)","Col4":"Total"}
        self.result['Profit and Loss by Month']['table1']['records'] = PL_json
        #---------Page 2 ----------------------------------
        head_json = {
            "customer": head.iloc[0,0],
            "report_type": head.iloc[1,0],
            "month_year":  head.iloc[2,0],
            "month1":  head.iloc[2,0].split('-')[0].strip(),
            "month2":  head.iloc[2,0].split('-')[1].split(',')[0].strip(),
            "no_tables":"1",
        }
        self.result['Page2'] ={}
        self.result['Page2']['header']=head_json
        dummy = self.pg2
        dummy.columns = ['Col1', 'Col2', 'Col3','Col4','Col5']
        PL_json = dummy.to_dict(orient='records')

        self.result['Page2']['table1']={}
        self.result['Page2']['table1']['name'] =''
        self.result['Page2']['table1']['no_cols'] ='5'
        self.result['Page2']['table1']['col-headers'] ={"Col1":"particulars","Col2":"Amount(Rs)","Col3":"% of total income","Col4":"Amount(Rs)","Col5":"% of total income"}
        self.result['Page2']['table1']['records'] = PL_json


        self.result = simplejson.dumps(self.result,ignore_nan=True)

        #self.result = simplejson.dumps(self.result)

#------------------Report 3 --------GST Summary-----------------------------------------
class Gst_Sum(RepGen):

    def __init__(self,file_name,file_name1=None,file_name2=None,file_suffix=None):
        super().__init__(file_name)

        self.df3 = pd.DataFrame(columns =['Mode Of Sales', 'Taxable values', 'CGST', 'SGST','Total GST'] )
        self.client_type1 =''
        self.client_type2 =''
        self.result={
            "customer":"abc",
            "month":"",
            "year":2019,
            "report_type":"GST",
            "b2c":[],
            "b2b":[],
            "input_tax_credit":0,
            "rcm":0,
            "total":{
                "tax_value":0,
                "cgst":0,
                "sgst":0,
                "total_gst":0
            },
            "no_of_days_delay":"",
            "interest":0,
            "late_payment_fees":0,
            "net_gst_payable":0

        }
        try:
            self.b2b = pd.read_excel(file_name1,sheet_name='b2b',skiprows=3)
            self.client_type2 ='B2B'
        except:
            self.status = 'Failure'
            raise
        try:
            self.b2c = pd.read_excel(file_name,skiprows=4)
            self.b2c.loc[self.b2c.iloc[:,4].str.contains('- EC', na=False, regex=True), 'Amount'] = 0
            self.client_type1 ='B2C'
            self.head=pd.read_excel(file_name,nrows=4,header=None)
        except:
            self.status = 'Failure'
            raise

        try:
            self.itc = pd.read_excel(file_name2,skiprows=4)
            self.itcbool =True
            self.head=pd.read_excel(file_name2,nrows=4,header=None)
        except:
            self.status = 'Failure'
            raise
        try:
            self.logo = 'fc_logo.png'
            self.Do_Accounting()
            self.Set_Writer(file_suffix)
            self.Is_Data_good()
            self.Set_Layout()
            self.Save_to_excel()
            #print('------------------Execution is complete---------------------')
            self.status = 'Success'
        except:
            self.status = 'Failure'
            raise

    def Set_Writer(self,file_suffix):
        self.fname='media/uploads/'+file_suffix+'GST_summary.xlsx'
        self.writer = pd.ExcelWriter(self.fname, engine='xlsxwriter')
        self.df3.to_excel(self.writer, index=False, sheet_name='GST',startrow=5,startcol=1)
        self.workbook = self.writer.book
        self.worksheet = self.writer.sheets['GST']

    def Do_Accounting(self):
        #-creating empty line
        self.df3 = self.df3.append({'Mode Of Sales':'', 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':''},ignore_index=True )


        #---Filter dataframe with string containing B2c and Total in mode of sales category----

        try:
            rp = self.b2c[pd.notnull(self.b2c.iloc[:,0])]
            r =rp[rp.iloc[:,0].str.contains('B2C')]
        except:
            pass
        b2ctot=0
        if self.client_type1 =='B2C' and r.empty is False:
            #----------
            indx = iter(r.index)
            valu=[]

            for ind in indx:
                val = self.b2c.iloc[ind:next(indx)-1]['Amount']
                val = val[val>=0]
                valu.append(val.sum())
                b2ctot = b2ctot+ val.sum()
            #---------------
            self.sales = r[r.iloc[:,0].str.contains('Total')]
            self.sales =self.sales[(self.sales['Amount']!='')]
            #--Create B2C line
            self.df3 = self.df3.append({'Mode Of Sales':'''B2C Sales (Sales to direct customers)''', 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':''},ignore_index=True )

            #----------Create dataframe with all existing mode of sales with GST report columns--
            itr=0
            for indx in self.sales.iloc[:,0].index:
                #mos = self.b2c.iloc[indx,0].split('Only B2C Sales')[0].split(' - ')[0].split('for')[1]
                mos = self.b2c.iloc[indx,0].split('Only B2C Sales')[0].split('for')[1]
                #tax_val = self.b2c.loc[indx,'Amount']
                tax_val= valu[itr]
                cgst = tax_val*0.025
                sgst = tax_val*0.025
                gst = cgst+sgst
                b2crecord={"name":mos,"tax_value":round(tax_val,2),"cgst":round(cgst,2),"sgst":round(sgst,2),"total_gst":round(gst,2)}
                self.result["b2c"].append(b2crecord)
                self.df3 = self.df3.append({'Mode Of Sales':mos, 'Taxable values':tax_val, 'CGST':cgst, 'SGST':sgst,'Total GST':gst},ignore_index=True)
                itr = itr+1
            #--Create Empty line
            self.df3 = self.df3.append({'Mode Of Sales':'', 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':''},ignore_index=True )

        #-----If B2B type client
        b2btot= 0
        if self.client_type2 =='B2B':
            #--Create B2B  line
            self.df3 = self.df3.append({'Mode Of Sales':'''B2B Sales (Sales to businesses)''', 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':''},ignore_index=True )

            #-- generate Taxable values sum by category of GST

            gst_sum = self.b2b.groupby('Rate')['Taxable Value'].sum()
            for items in gst_sum.index.values:
                mos = str(int(items)) + '% GST category'
                tax_val = gst_sum[[items]].values[0]
                cgst = tax_val*items/100/2
                sgst = tax_val*items/100/2
                gst = cgst+sgst
                b2btot = b2btot+ tax_val
                b2brecord={"name":mos,"tax_value":round(tax_val,2),"cgst":round(cgst,2),"sgst":round(sgst,2),"total_gst":round(gst,2)}
                self.result["b2b"].append(b2brecord)
                self.df3 = self.df3.append({'Mode Of Sales':mos, 'Taxable values':tax_val, 'CGST':cgst, 'SGST':sgst,'Total GST':gst},ignore_index=True)
            #------------------ITC-------------
            try:
                r =self.itc[ pd.notnull(self.itc['Description'])]
                p =r[r['Description'].str.contains('ITC Available')]
                itc = p[['Integrated Tax','Central Tax','State/UT Tax']].sum().sum()
                itc = itc*(b2btot/(b2btot+ b2ctot))
                #print (itc,b2btot,b2ctot)
            except:
                itc = 'NA'

            mos = 'input tax credit'
            self.result["input_tax_credit"]=round(itc,2)
            self.df3 = self.df3.append({'Mode Of Sales':mos, 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':itc},ignore_index=True)
            #------------------RCM-------------
            rcm = 0
            mos = 'RCM'
            self.result["rcm"]=round(rcm,2)
            self.df3 = self.df3.append({'Mode Of Sales':mos, 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':rcm},ignore_index=True)


        #--Summing up values to a row---
        gst_sales  = self.df3[(self.df3['Taxable values']!='')]

        tax_val = gst_sales['Taxable values'].sum()
        cgst = gst_sales['CGST'].sum()
        sgst = gst_sales['SGST'].sum()
        gst = gst_sales['Total GST'].sum()
        self.result["total"]={
            "tax_value":round(tax_val,2),
            "cgst":round(cgst,2),
            "sgst":round(sgst,2),
            "total_gst":round(gst,2)
        }
        self.df3 = self.df3.append({'Mode Of Sales':'Total', 'Taxable values':tax_val, 'CGST':cgst, 'SGST':sgst,'Total GST':gst},ignore_index=True)

        #--------Filing delay and other values--------------------
        mon = self.head.iloc[2,0].split(' ')[0]
        month = strptime(mon,'%B').tm_mon
        year = int(self.head.iloc[2,0].split(' ')[1])
        # print(mon)
        customer1 = self.head.iloc[0,0]
        self.result["month"]=mon
        self.result["year"]=year
        self.result["customer"]=str(customer1).lower().capitalize()
        delay = (datetime.today() - (datetime(year,month,20) + relativedelta(months=+1))).days
        if delay < 0:
            delay=0
        self.df3 = self.df3.append({'Mode Of Sales':'No. of days of delay in filing', 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':str(delay)+' days'},ignore_index=True )
        self.result["no_of_days_delay"]=str(delay)+' days'
        #-----Interest---------------------
        interest = gst*delay*0.18/365
        self.result["interest"]=round(interest,2)
        self.df3 = self.df3.append({'Mode Of Sales':'Interest', 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':interest},ignore_index=True )

        #-----Late filing fees---------------------
        fee = 50*delay
        self.df3 = self.df3.append({'Mode Of Sales':'Late filing fees', 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':fee},ignore_index=True )
        self.result["late_payment_fees"]=round(fee,2)
        #-----Net GST-------------------------------
        netgst = gst+ interest + fee
        self.df3 = self.df3.append({'Mode Of Sales':'NET GST PAYABLE', 'Taxable values':'', 'CGST':'', 'SGST':'','Total GST':netgst},ignore_index=True )
        self.result["net_gst_payable"]=round(netgst,2)

    def Set_Layout(self):
        #hide_gridlines(self.worksheet,2)
        fmt= self.workbook.add_format({'num_format': '#,##0.00 _€'})
        self.worksheet.set_column('B:B', 30,fmt)
        self.worksheet.set_column('C:F', 20,fmt)
        self.worksheet.set_column('G:G', 30,fmt)
        self.worksheet.set_column('A:A', 4,fmt)

        cell = 'A'+ str(self.df3.shape[0]+ 6)
        insert_image(self.worksheet,self.logo,cell)
        try:
            merge_cells(self.workbook,self.worksheet,'A1','F1', str(self.head.iloc[0,0]))
            merge_cells(self.workbook,self.worksheet,'A2','F2', 'GST computation & summary')
            merge_cells(self.workbook,self.worksheet,'A3','F3', str(self.head.iloc[2,0]))
        except:
            raise




    def Save_to_excel(self):
        self.writer.save()

#-------------Report 6 --Closing Stock--------------------------
class Closing_Stock(RepGen):

    def __init__(self,file_name1,file_name2,file_suffix=None):
        super().__init__(file_name1)
        try:
            print("Here1")
            print(file_name1)
            self.weight = pd.read_excel(file_name1,usecols="L",nrows=3)
            print("here4")
            if self.weight.iloc[:,0].str.contains('Weighted Average').any():
                print("Hey here")
                self.L2 = 'Weighted Average'
            elif self.weight.iloc[:,0].str.contains('Last Invoice Purchase Rate').any():
                print("Here3")
                self.L2 = 'Last Invoice Purchase Rate'

            print("Here2")
            self.Alert = pd.DataFrame(columns =['SI.No','Item Name','Alert type','Alert on basis','Value','Deviation %', 'Benchmark Value','Date','No.'])
            self.df2 = pd.read_excel(file_name1,usecols="A:D")
            self.price = pd.read_excel(file_name2,skiprows=4)
            self.head=pd.read_excel(file_name2,nrows=4,header=None)


            self.Do_Accounting()
            print("Accounting done")
            self.Set_Writer(file_suffix)
            print("Writer done")
            self.Is_Data_good()
            print("data gppd done")
            self.Set_Layout()
            print("layout done")
            self.Save_to_excel()
            print("save to excel done")
            self.status = 'Success'
        except:
            self.status = 'Failure'
            raise

    def Set_Writer(self,file_suffix):
        self.fname = 'media/uploads/'+file_suffix+slugify.slugify(str(self.head.iloc[2,0].split('-')[1]).strip())+'_Closing_stock_report.xlsx'
        self.fname=self.fname.strip()
        self.writer = pd.ExcelWriter(self.fname, engine='xlsxwriter')
        self.fname1 = 'media/uploads/'+file_suffix+'Alert.xlsx'
        self.Alertwriter = pd.ExcelWriter(self.fname1, engine='xlsxwriter')
        self.df2.to_excel(self.writer, index=False, sheet_name='Closing stock')
        self.Alert.to_excel(self.Alertwriter, index=False,sheet_name='Alerts')
        self.workbook = self.writer.book
        self.Alertbook = self.Alertwriter.book
        self.worksheet = self.writer.sheets['Closing stock']
        self.Alertsheet= self.Alertwriter.sheets['Alerts']

    def Do_Accounting(self):
        print("Accounting")
        self.df2['Price']=np.nan
        self.df2['Total']=np.nan
        count =1
        for items in self.df2['Item Name']:
            print(items)
            trans_list = self.price[self.price['Product/Service']==items]
            print(trans_list)
            if not trans_list.empty:
                trans_list.reset_index(inplace=True)

                mode_rate = trans_list['Rate'].mode()[0]
                mean_rate = trans_list['Rate'].mean()
                mean_qty = trans_list['Qty'].mean()
                sum_qty = trans_list['Qty'].sum()

                for i,rows in trans_list.iterrows():
                    if float(trans_list.loc[i,'Rate']) > 1.2*mode_rate:
                        dev = 100*(trans_list.loc[i,'Rate'] - mode_rate)/mode_rate
                        self.Raise_Alert(count,items,'Rate_Mode','trans > 1.20*Mode',trans_list.loc[i,'Rate'],dev,mode_rate,trans_list.loc[i,'Date'],trans_list.loc[i,'No.'])
                        count = count +1
                        #print (items,'Mode-Alert-the transaction',trans_list.loc[i,'Rate'],' is 1.2 times >',mode_rate)
                    if float(trans_list.loc[i,'Rate']) > 1.2*mean_rate:
                        dev = 100*(trans_list.loc[i,'Rate'] - mean_rate)/mean_rate
                        self.Raise_Alert(count,items,'Rate_Mean','trans > 1.20*Mean',trans_list.loc[i,'Rate'],dev,mean_rate,trans_list.loc[i,'Date'],trans_list.loc[i,'No.'])
                        count = count +1
                        #print (items,'Mean-Alert-the transaction',trans_list.loc[i,'Rate'],'is 1.2 times >',mean_rate)
                    """if float(trans_list.loc[i,'Qty']) > 1.2*mean_qty:
                        dev = 100*(trans_list.loc[i,'Qty'] - mean_qty)/mean_qty
                        self.Raise_Alert(count,items,'Qty_Mean','trans > 1.20*Mean',trans_list.loc[i,'Qty'],dev,mean_qty,trans_list.loc[i,'Date'],trans_list.loc[i,'No.'])
                        count = count +1"""
                    #print (items,'Mode-Alert-the qantity',trans_list.loc[i,'Qty'],' is 1.2 times >',mode_qty)
                    if float(trans_list.loc[i,'Qty']) > 0.25*sum_qty:
                        dev = 100*(trans_list.loc[i,'Qty'])/sum_qty
                        self.Raise_Alert(count,items,'Qty_sum','trans > 1.25*sum',trans_list.loc[i,'Qty'],dev,sum_qty,trans_list.loc[i,'Date'],trans_list.loc[i,'No.'])
                        count = count +1
                        #print (items,'Mean-Alert-the quantity',trans_list.loc[i,'Qty'],'is 1.2 times >',mean_qty)

                stock = float(self.df2[self.df2['Item Name']==items]['Stock in Store'])
                if stock > 0.30*(sum_qty/3.0):
                    dev = 100*stock/sum_qty/3.0
                    self.Raise_Alert(count,items,'stock in store','Stock > 30% of montly avg',stock,dev,sum_qty/3.0,trans_list.loc[i,'Date'],trans_list.loc[i,'No.'])
                    count = count +1
                if self.L2 == 'Weighted Average':
                    avgrate = self.price[self.price['Product/Service']==items]['Amount'].sum()
                    qty = self.price[self.price['Product/Service']==items]['Qty'].sum()
                    self.df2.loc[self.df2['Item Name']== items,'Price']= avgrate/qty
                elif self.L2 == 'Last Invoice Purchase Rate':
                    rate = self.price[self.price['Product/Service']==items]['Rate'].iloc[-1]
                    self.df2.loc[self.df2['Item Name']== items,'Price']= rate


        print("After for loop")
        for i, rows in self.df2.iterrows():
            if pd.notna(self.df2.loc[i,'Stock in Store']) and pd.notna(self.df2.loc[i,'Price']):
                self.df2.loc[i,'Total'] = float(self.df2.loc[i,'Stock in Store'])*float(self.df2.loc[i,'Price'])

        ind =self.df2.shape[0]
        self.df2.loc[ind+1,'Total'] = self.df2['Total'].sum()


        self.df2.update(self.df2.round(2))
        self.Alert.update(self.Alert.round(2))

    def Set_Layout(self):
        #hide_gridlines(self.worksheet,2)
        fm = self.Alertbook.add_format({'align': 'center'})
        self.worksheet.set_column('A:A', 5)
        self.worksheet.set_column('B:B', 30)
        self.worksheet.set_column('C:C', 10)
        self.worksheet.set_column('D:D', 20)
        self.worksheet.set_column('E:F', 10)

        self.Alertsheet.set_column('A:A', 5)
        self.Alertsheet.set_column('B:G', 20)
        self.Alertsheet.set_column('H:H', 20,fm)
        self.Alertsheet.set_column('H:I', 20,fm)

        lst = self.df2.index[self.df2['S.No'].isnull() & self.df2['Stock in Store'].isnull() &  self.df2['Item Name'].notnull() &  self.df2['UOM'].isnull()].to_list()
        fmt = self.workbook.add_format({'bold': True})
        for ind in lst:
            self.worksheet.write(ind+1,1, self.df2.iloc[ind,1], fmt)


    def Save_to_excel(self):
        self.writer.save()
        self.Alertwriter.save()

    def Raise_Alert(self,sno,inm,at,ab,val,dev,bv,dt,billno):
        self.Alert = self.Alert.append({
            'SI.No':sno,
            'Item Name':inm,
            'Alert type':at,
            'Alert on basis':ab,
            'Value':val,
            'Deviation %':dev,
            'Benchmark Value':bv,
            'Date':dt,
            'No.':billno
        },ignore_index=True)

#-------------Report 6 --Purchase Efficiency--------------------------
class Purchase_Efficiency(RepGen):
    def __init__(self,file_name,file_suffix=None):
        super().__init__(file_name)
        try:
            self.dfs = pd.read_excel(file_name, header=None)
            self.result =''
            self.rep_create(file_name,file_suffix)
            print("Before writer")
            self.Set_Writer()
            print("Hey whats up?")
            self.Set_Layout()
            self.Save_to_excel()
            self.status = 'Success'
        #print('------------------Execution is complete---------------------')
        except:
            self.status = 'Failure'
            raise

    def Set_Writer(self):
        self.workbook = self.writer.book
        self.worksheet = self.writer.sheets['Itemised details']
        self.page2 = self.writer.sheets['Summary']

    def rep_create(self, file_name,file_suffix):
        try:
            #print(file_suffix)
            self.df1 = pd.read_excel(file_name, skiprows=4)
            self.fname = 'media/uploads/' + file_suffix + 'Purchase_Efficiency_Report.xlsx'
            self.writer = pd.ExcelWriter(self.fname, engine='xlsxwriter')

            self.df1 = self.df1.dropna(subset=['Product/Service'])
            self.df1 = self.df1.drop(
                columns=['Unnamed: 0', 'Memo/Description', 'HSN/SAC', 'Balance', 'Transaction Type'])
            self.df1 = self.df1.rename(
                columns={'Product/Service': 'ItemName', 'No.': 'BillNo', 'Supplier': 'Vendor', 'Qty': 'Quantity',
                         'Rate': 'PurchasePrice', \
                         'Amount': 'ActualPurchasingCost'})
            self.df1['IdealPrice'] = (self.df1.groupby(['ItemName'])['PurchasePrice'].transform('min'))
            self.df1['ExtraPricePaid'] = self.df1['PurchasePrice'] - self.df1['IdealPrice']
            self.df1['IdealPurchasingCost'] = self.df1['IdealPrice'] * self.df1['Quantity']
            self.df1['EfficiencyLoss'] = self.df1['ExtraPricePaid'] * self.df1['Quantity']
            self.df1 = self.df1[self.df1['EfficiencyLoss'] != 0]
            self.df1 = self.df1[
                ['ItemName', 'Date', 'BillNo', 'Vendor', 'PurchasePrice', 'IdealPrice', 'ExtraPricePaid', 'Quantity',
                 'EfficiencyLoss', 'ActualPurchasingCost', \
                 'IdealPurchasingCost']]

            # Item Summary Table
            self.df1_itempivot = pd.pivot_table(self.df1, values=['Quantity', 'PurchasePrice', 'IdealPrice',
                                                                  'ActualPurchasingCost', 'IdealPurchasingCost',
                                                                  'EfficiencyLoss'], \
                                                index=['ItemName'],
                                                aggfunc={'Quantity': np.sum, 'PurchasePrice': np.mean,
                                                         'IdealPrice': np.mean, \
                                                         'ActualPurchasingCost': np.sum, 'IdealPurchasingCost': np.sum,
                                                         'EfficiencyLoss': np.sum})
            self.df1_itempivot = self.df1_itempivot[
                ['Quantity', 'PurchasePrice', 'IdealPrice', 'ActualPurchasingCost', 'IdealPurchasingCost',
                 'EfficiencyLoss']]
            self.df1_itempivot = self.df1_itempivot.sort_values(by='EfficiencyLoss', ascending=False)
            self.df1_itemsumry = pd.DataFrame(self.df1_itempivot.to_records())
            #         display(df1_itemsumry.head())
            self.df1_itemsumry = self.df1_itemsumry.head()

            # Date Summary Table
            self.df1_datepivot = pd.pivot_table(self.df1, values=['ActualPurchasingCost', 'IdealPurchasingCost',
                                                                  'EfficiencyLoss'], index=['Date'], \
                                                aggfunc={'ActualPurchasingCost': np.sum, 'IdealPurchasingCost': np.sum,
                                                         'EfficiencyLoss': np.sum})
            self.df1_datepivot = self.df1_datepivot[['ActualPurchasingCost', 'IdealPurchasingCost', 'EfficiencyLoss']]
            self.df1_datepivot = self.df1_datepivot.sort_values(by='EfficiencyLoss', ascending=False)
            self.df1_datesumry = pd.DataFrame(self.df1_datepivot.to_records())
            #         display(df1_datesumry.head())
            self.df1_datesumry = self.df1_datesumry.head()

            # Vendor Summary Table
            self.df1_vendorpivot = pd.pivot_table(self.df1, values=['ActualPurchasingCost', 'IdealPurchasingCost',
                                                                    'EfficiencyLoss'], index=['Vendor'], \
                                                  aggfunc={'ActualPurchasingCost': np.sum,
                                                           'IdealPurchasingCost': np.sum, 'EfficiencyLoss': np.sum})
            self.df1_vendorpivot = self.df1_vendorpivot[
                ['ActualPurchasingCost', 'IdealPurchasingCost', 'EfficiencyLoss']]
            self.df1_vendorpivot = self.df1_vendorpivot.sort_values(by='EfficiencyLoss', ascending=False)
            self.df1_vendorsumry = pd.DataFrame(self.df1_vendorpivot.to_records())
            #         display(df1_vendorsumry.head())
            self.df1_vendorsumry = self.df1_vendorsumry.head()

            # bill details table
            self.df_detailed = self.df1.drop(columns=['ActualPurchasingCost', 'IdealPurchasingCost'])
            cols = ["ExtraPricePaid", "Quantity", "EfficiencyLoss"]
            self.df_detailed['ItemName'] = self.df_detailed['ItemName'].ffill()

            grand = self.df_detailed[cols].sum()
            grand.loc['ItemName'] = 'Grand total'
            #         print (grand)

            self.df1_detailed = self.df_detailed.groupby('ItemName')[cols].sum()
            self.df1_detailed.index = self.df1_detailed.index + '____'

            # create empty DataFrame
            self.df2_detailed = pd.DataFrame(index=self.df1_detailed.index + '__')
            self.df_detailed = pd.concat([self.df_detailed.set_index('ItemName'), self.df1_detailed, self.df2_detailed],
                                         keys=('a', 'b', 'c')).sort_index(level=1).reset_index()

            # get output by 2 conditions
            m1 = self.df_detailed['level_0'] == 'a'
            m2 = self.df_detailed['level_0'] == 'c'
            self.df_detailed['ItemName'] = np.select([m1, m2], [self.df_detailed['ItemName'], np.nan], default='Total')
            self.df_detailed = self.df_detailed.drop('level_0', axis=1)
            self.df_detailed.loc[len(self.df_detailed.index)] = grand
            self.df_detailed = self.df_detailed.dropna(subset=['ItemName'])
            self.df_detailed = self.df_detailed[
                ['ItemName', 'Date', 'BillNo', 'Vendor', 'PurchasePrice', 'IdealPrice', 'ExtraPricePaid', 'Quantity',
                 'EfficiencyLoss']]
            self.df_detailed.fillna('', inplace=True)

            L1 = []
            for item in self.df1_itemsumry.iterrows():
                itemDetails = self.df_detailed[self.df_detailed['ItemName'] == item[1][0]].to_json(orient="records")
                L1.append(itemDetails)
            self.df1_itemsumry.insert(7, "itemDetails", L1, True)

            L2 = []
            for i in self.df1_datesumry.iterrows():
                details = self.df_detailed[self.df_detailed['Date'] == i[1][0]].to_json(orient="records")
                L2.append(details)
            self.df1_datesumry.insert(4, "itemDetails", L2, True)

            L3 = []
            for j in self.df1_vendorsumry.iterrows():
                vDetails = self.df_detailed[self.df_detailed['Vendor'] == j[1][0]].to_json(orient="records")
                L3.append(vDetails)
            self.df1_vendorsumry.insert(4, "itemDetails", L3, True)

            result_itemsummary = self.df1_itemsumry.to_json(orient="records")
            result_datesummary = self.df1_datesumry.to_json(orient="records")
            result_vendorsummary = self.df1_vendorsumry.to_json(orient="records")
            result_billdetailed = self.df_detailed.to_json(orient="records")
            self.df1_itemsumry = self.df1_itemsumry.drop(columns=['itemDetails'])
            self.df1_datesumry = self.df1_datesumry.drop(columns=['itemDetails'])
            self.df1_vendorsumry = self.df1_vendorsumry.drop(columns=['itemDetails'])
            self.df1_itemsumry.to_excel(self.writer, index=False, sheet_name='Summary', startrow=4)
            self.df1_datesumry.to_excel(self.writer, index=False, sheet_name='Summary', startrow=13)
            self.df1_vendorsumry.to_excel(self.writer, index=False, sheet_name='Summary', startrow=22)
            self.df_detailed.to_excel(self.writer, index=False, sheet_name='Itemised details', startrow=4)

            #         parsed4 = json.loads(result_billdetailed)
            header = pd.read_excel(file_name,nrows=4,header=None)
            header = header[[0]]
            client_name = header[0][0]
            month_yr = header[0][2]
            # m = re.search('- (.+?),', month_yr)
            m = month_yr[0:3]
            y = month_yr[-4:]
            # if m:
            #     found = m.group(1)
            # mn_yr = found[:3] + ' ' +y
            mn_yr = m + ' ' +y
            rep_name = "Purchase Efficiency Report"
            head_json = {"customer":client_name, "report_type":rep_name, "month_year":mn_yr }

            # self.writer = pd.ExcelWriter('Purchase Efficiency Report.xlsx', engine='xlsxwriter')
            self.result={"head": head_json, "itemsummaryTable": json.loads(result_itemsummary), "datesummaryTable": json.loads(result_datesummary), "vendorsummaryTable": json.loads(result_vendorsummary), "billdetailedTable": json.loads(result_billdetailed) }
            self.result=simplejson.dumps(self.result,ignore_nan=True)


        except Exception as e:
            self.status = 'Failure'
            self.pass_error(e)
            self.errors="Function rep_create failed executing correctly"

    def Set_Layout(self):

        merge_cells(self.workbook,self.worksheet,'A1','I1', str(self.dfs.iloc[0,0]))
        merge_cells(self.workbook,self.worksheet,'A2','I2', 'Itemised Details')
        merge_cells(self.workbook,self.worksheet,'A3','I3', str(self.dfs.iloc[2,0]))

        merge_cells(self.workbook,self.page2,'A1','G1', str(self.dfs.iloc[0,0]))
        merge_cells(self.workbook,self.page2,'A2','G2', 'Purchase Efficiency Summary')
        merge_cells(self.workbook,self.page2,'A3','G3', str(self.dfs.iloc[2,0]))


    def Save_to_excel(self):
        self.writer.save()

#--------------Report 7 -------Swiggy Dump Level Reconciliation--------------------
class SDLR(RepGen):

    def __init__(self,path1,path2,file_name2,prefix):
        super().__init__(file_name2)

        self.master = pd.read_excel(file_name2)
        print("Hey called SDLR")
        print(path1)
        self.path1 = path1
        self.path2 = path2
        self.cal = pd.DataFrame()
        self.status="Failure"
        self.fname1=""
        self.fname2=""
        #self.outlet_del = pd.DataFrame(columns=['ID','Date', 'Order Id', 'Price (As per Swiggy)', 'Price (As per Management)'])
        #self.outlet_can = pd.DataFrame(columns=['ID','Date', 'Order Id', 'Price (As per Swiggy)', 'Price (As per Management)'])
        #self.outlet_price = pd.DataFrame(columns=['ID','Date', 'Order Id', 'Price (As per Swiggy)', 'Price (As per Management)'])

        self.swiggy1 = pd.DataFrame(columns=['S. No', 'Particulars', 'No. Of Orders', 'Value of orders','Impact on cash flow'])
        self.swiggy2 = pd.DataFrame(columns=['S. No', 'Particulars', 'No. Of Orders', 'Value of orders as per Swiggy pricing','Value of orders as per Management pricing','Impact on cash flow'])
        try:
            self.Load_Dumps()
            f = open("sdump_log.txt", "a")
            f.write("Load Dumps Completed\n")
            f.close()
            self.Load_Mail()
            f = open("sdump_log.txt", "a")
            f.write("Load Mail Completed\n")
            f.close()
            self.Do_Accounting()
            f = open("sdump_log.txt", "a")
            f.write("Do accounting Completed\n")
            f.close()
            self.Set_Writer(prefix)
            f = open("sdump_log.txt", "a")
            f.write("Set writer Completed\n")
            f.close()
            self.Is_Data_good()
            f = open("sdump_log.txt", "a")
            f.write("Is data good Completed\n")
            f.close()
            self.Set_Layout()
            f = open("sdump_log.txt", "a")
            f.write("Set Layout Completed\n")
            f.close()
            self.Save_to_excel()
            f = open("sdump_log.txt", "a")
            f.write("Saved to excel\n")
            f.close()
            print("saved to excel")
            print('------------------Execution is complete---------------------')
            self.status="Success"
        except:
            self.status="Failure"

    def Load_Dumps(self):
        print("load dumps")
        all_files = glob.glob(self.path1 + "*.csv")


        fi = []
        li = []
        for filename in all_files:

            df = pd.read_csv(filename,skiprows=5,names= range(0,55))
            dfst = pd.read_csv(filename,skiprows=5,usecols=range(0,30))
            loc = dfst.columns.get_loc("MOU type")
            head = pd.read_csv(filename,nrows=3)
            outletid = head.index[2][1]
            print(loc,outletid)
            print(type(df.iloc[:,loc]))
            df.iloc[:,loc] = str(outletid) +'_'+ df.iloc[:,loc].astype(str)
            dfst.iloc[:,loc] = str(outletid) +'_'+ dfst.iloc[:,loc].astype(str)
            df=  df.drop(df.index[0])
            fi.append(df)
            li.append(dfst)

        self.dump = pd.concat(fi, axis=0, ignore_index=True)
        self.dfst = pd.concat(li, axis=0, ignore_index=True)

        print('Completed loading dump files -', self.dump.shape)

    def Load_Mail(self):
        all_files = glob.glob(self.path2 + "*.xlsx")
        fi = []

        for filename in all_files:
            df = pd.read_excel(filename)
            fi.append(df)

        self.mail = pd.concat(fi, axis=0, ignore_index=True)
        print('Completed loading Mail files- ',self.mail.shape)

    def Do_Accounting(self):
        #self.dump = self.dump.rename(columns=self.dump.iloc[0])

        loc = self.dfst.columns.get_loc("Restaurant-bear")+1
        mou = self.dfst.columns.get_loc("MOU type")
        print ('column location-',loc,' mou-',mou)
        #self.dump = self.dump.dropna(axis =1, how = 'all',subset=range(loc,self.dump.shape[1]))

        self.dump = self.dump.reset_index(drop=True)
        self.dfst = self.dfst.reset_index(drop=True)



        data = pd.DataFrame()
        for i in range(loc,self.dump.shape[1]):
            print (i, self.dump.shape[1])
            cal = self.dump[self.dump.columns[i]]
            #print(self.dump.iloc[:,i].isnull().all())
            try:
                data = self.dump.iloc[:,i].str.split('_',expand=True)
                data['apm'] = list(map(float,(self.Get_Mgt_Val(data))))
                data[3] = data[3].str.split('+',expand=True)[0].astype(float)
                self.cal =pd.concat([self.cal,cal,data],axis=1)
            except:
                print('empty column-',i)
        f = open("sdump_log.txt", "a")
        f.write("Splitting columns completed\n")
        f.close()
        self.cal['Total Value Mngmt'] = self.cal['apm'].sum(axis=1)
        self.cal['Final Order Value Mngmt'] = self.cal['Total Value Mngmt']+ self.dump[21].astype(float)
        self.cal['Swiggy Total'] =self.cal[3].sum(axis=1)
        self.cal['diff'] = self.cal['Final Order Value Mngmt'] - self.cal['Swiggy Total']

        self.cal = pd.concat([self.dump[self.dump.columns[0:loc]],self.cal],axis=1)
        f = open("sdump_log.txt", "a")
        f.write("Completed ceating cal dataframe\n")
        f.close()
        #--------------------Missing Data--------------------------------------
        for ind in self.dump[self.dump.columns[0]].index:
            item = self.dump.iloc[ind,0]
            if np.mod(ind,1000)==0:
                print (ind,'of data compared from total of',self.dump.shape[0])
            if  not self.mail[self.mail['Order No']==int(item)].empty:
                self.cal.loc[ind,'invoice_missing'] = 'No'
            else:
                self.cal.loc[ind,'invoice_missing'] = 'Yes'
        print('Completed listing missing data')

        self.cal.iloc[:,mou] = self.cal.iloc[:,mou].str.split('_',expand=True)[0]
        #---------delivered-----------------------
        r = self.cal[self.cal['invoice_missing']=='Yes']
        r = r[r.iloc[:,1]=='delivered']
        p = r.iloc[:,[mou,2,0]]
        k = r[['Final Order Value Mngmt','Swiggy Total']]
        self.outlet_del = pd.concat([p,k],axis=1)
        self.outlet_del.columns = ['ID','Date','Order Id','Price (As per Swiggy)','Price (As per Management)']

        num_ord_del = r.shape[0]
        val_ord_del = r['Final Order Value Mngmt'].sum()
        print('Completed creating missinginvoice for delivered data ')
        #---------cancelled-----------------------
        r = self.cal[self.cal['invoice_missing']=='Yes']
        r = r[r.iloc[:,1]=='cancelled']
        p = r.iloc[:,[mou,2,0]]
        k = r[['Final Order Value Mngmt','Swiggy Total']]
        self.outlet_can = pd.concat([p,k],axis=1)
        self.outlet_can.columns = ['ID','Date','Order Id','Price (As per Swiggy)','Price (As per Management)']


        num_ord_can = r.shape[0]
        val_ord_can = r['Final Order Value Mngmt'].sum()
        print('Completed creating missinginvoice for cancelled data ')
        #-----------------pricing mismatch---------------------
        #---------higher-----------------------

        r = self.cal[self.cal['diff']!=0.0]
        r = self.cal[self.cal['invoice_missing']=='No']
        r = r[r.iloc[:,1]=='delivered']
        r = r[r['diff']>0]
        #-------eliminating missing from master list
        r = r[r['diff']<10000]
        p = r.iloc[:,[mou,2,0]]
        k = r[['Final Order Value Mngmt','Swiggy Total']]
        self.outlet_price1 = pd.concat([p,k],axis=1)

        num_ord_del_s = r.shape[0]
        val_ord_del_s = r['Final Order Value Mngmt'].sum()
        val_swiggy_del_s = r['Swiggy Total'].sum()

        #---------lower-----------------------

        r = self.cal[self.cal['diff']!=0.0]
        r = self.cal[self.cal['invoice_missing']=='No']
        r = r[r.iloc[:,1]=='delivered']
        r = r[r['diff']<0]
        p = r.iloc[:,[mou,2,0]]
        k = r[['Final Order Value Mngmt','Swiggy Total']]
        self.outlet_price2 = pd.concat([p,k],axis=1)


        num_ord_can_s = r.shape[0]
        val_ord_can_s = r['Final Order Value Mngmt'].sum()
        val_swiggy_can_s = r['Swiggy Total'].sum()
        self.outlet_price = pd.concat([self.outlet_price1,self.outlet_price2],axis=0,ignore_index=True)
        self.outlet_price.columns = ['ID','Date','Order Id','Price (As per Swiggy)','Price (As per Management)']

        tot_ord = num_ord_del_s+ num_ord_can_s
        tot_swiggy = val_swiggy_del_s+val_swiggy_can_s
        tot_val = val_ord_can_s+val_ord_del_s

        print('Completed creating list pricing data data ')
        #-----------------Swiggy summary sheet---------------------------
        self.swiggy1 = self.swiggy1.append({'S. No':'', 'Particulars':'', 'No. Of Orders':'', 'Value of orders':'','Impact on cash flow':''},ignore_index=True )
        self.swiggy1 = self.swiggy1.append({'S. No':'', 'Particulars':'', 'No. Of Orders':'', 'Value of orders':'','Impact on cash flow':''},ignore_index=True )

        self.swiggy1 = self.swiggy1.append({'S. No':'A.', 'Particulars':'Revenue impact: Difference in orders (Delivered)', 'No. Of Orders':'', 'Value of orders':'','Impact on cash flow':''},ignore_index=True )
        self.swiggy1 = self.swiggy1.append({'S. No':'', 'Particulars':'i. Orders recorded in the dump, not available in the invoices', 'No. Of Orders':num_ord_del, 'Value of orders':val_ord_del,'Impact on cash flow':''},ignore_index=True )

        self.swiggy1 = self.swiggy1.append({'S. No':'B.', 'Particulars':'Revenue impact: Difference in orders (Cancelled)', 'No. Of Orders':num_ord_can, 'Value of orders':val_ord_can,'Impact on cash flow':'''Can't be assessed at this level. Parametres like reasons for cancellation, cancellation charges and conditions are to be known.'''},ignore_index=True )

        #------------------------------
        self.swiggy2 = self.swiggy2.append({'S. No':'A', 'Particulars':'Revenue impact: Difference in pricing (Delivered)', 'No. Of Orders':'', 'Value of orders as per Swiggy pricing':'','Value of orders as per Management pricing':'','Impact on cash flow':''},ignore_index=True )
        self.swiggy2 = self.swiggy2.append({'S. No':'', 'Particulars':'i. Orders for which lesser price is charged by Swiggy', 'No. Of Orders':num_ord_del_s, 'Value of orders as per Swiggy pricing':val_swiggy_del_s,'Value of orders as per Management pricing':val_ord_del_s,'Impact on cash flow':''},ignore_index=True )
        self.swiggy2 = self.swiggy2.append({'S. No':'', 'Particulars':'ii. Orders for which a higher price is charged by Swiggy', 'No. Of Orders':num_ord_can_s, 'Value of orders as per Swiggy pricing':val_swiggy_can_s,'Value of orders as per Management pricing':val_ord_can_s,'Impact on cash flow':''},ignore_index=True )
        self.swiggy2 = self.swiggy2.append({'S. No':'', 'Particulars':'Total', 'No. Of Orders':tot_ord, 'Value of orders as per Swiggy pricing':'','Value of orders as per Management pricing':'','Impact on cash flow':''},ignore_index=True )


    def Get_Mgt_Val(self,dat):
        lst =[]
        for ind in dat.index:
            try:
                if '+' in dat.iloc[ind,3]:
                    price = self.master[self.master['ItemName']==dat.iloc[ind,0]]['As per Managament'].values[0]
                    addon =self.master[self.master['ItemName']==dat.iloc[ind,0]+'+'+dat.iloc[ind,3].split('+')[1]]['As per Managament'].values[0]
                    val =(float(price)+float(addon))*float(dat.iloc[ind,2])
                    lst.append(val)
                else:
                    price = self.master[self.master['ItemName']==dat.iloc[ind,0]]['As per Managament'].values[0]
                    val = price*float(dat.iloc[ind,2])
                    lst.append(val)

            except:
                if type(dat.iloc[ind,0]) ==str:
                    if '+' in dat.iloc[ind,3]:
                        a = dat.iloc[ind,3].split('+')[0]
                        lst.append(float(a))
                    else:
                        a = dat.iloc[ind,3]
                        lst.append(float(a))
                else:
                    lst.append(np.nan)
        return lst

    def Set_Writer(self,prefix):
        output_filepath1=prefix+'SDLR.xlsx'
        output_filepath2=prefix+'SDLR_summary.xlsx'
        self.fname1=output_filepath1
        self.fname2=output_filepath2
        self.writer = pd.ExcelWriter(output_filepath1, engine='xlsxwriter')
        self.swiggywriter = pd.ExcelWriter(output_filepath2, engine='xlsxwriter')
        self.cal.to_excel(self.writer, index=False, sheet_name='SDLR')
        self.swiggy1.to_excel(self.swiggywriter, index=False, sheet_name='Summary',startrow = 3,startcol=1)
        self.swiggy2.to_excel(self.swiggywriter, index=False, sheet_name='Summary',startrow = 3+ self.swiggy1.shape[0]+3)

        self.summarysheet = self.swiggywriter.sheets['Summary']

        self.outlet_del.to_excel(self.swiggywriter, index=False, sheet_name='Detailed report',startrow = 3)
        self.outlet_can.to_excel(self.swiggywriter, index=False, sheet_name='Detailed report',startrow = 3+self.outlet_del.shape[0]+5)
        self.outlet_price.to_excel(self.swiggywriter, index=False, sheet_name='Detailed report',startrow = 3+ self.outlet_del.shape[0]+ self.outlet_can.shape[0]+3+5)
        self.detailsheet = self.swiggywriter.sheets['Detailed report']
        self.detailsheet.write_string(2+ self.outlet_del.shape[0]+ self.outlet_can.shape[0]+3+5, 0, ' Differences in pricing')
        self.detailsheet.write_string(2+self.outlet_del.shape[0]+5, 0, ' Differences in cancelled orders')
        self.detailsheet.write_string(2, 0, ' Differences in delivered orders')


        self.workbook = self.writer.book
        self.worksheet = self.writer.sheets['SDLR']
        self.swiggybook = self.swiggywriter.book

    def Is_Data_good(self):
        pass

    def Set_Layout(self):
        fmt = self.swiggybook.add_format({'text_wrap': True})
        self.summarysheet.set_column('B:F', 35,fmt)
        self.detailsheet.set_column('A:F', 20,fmt)

    def Save_to_excel(self):
        self.writer.save()
        self.swiggywriter.save()

#-------------Report 6 --Swiggy Invoice level Reconciliation--------------------------
class SILR(RepGen):
    def __init__(self,path1,path2):
        try:
            super().__init__(path1)
            self.path1 = path1
            self.path2 = path2
            self.make_copy()
            self.Do_Accounting()
            self.Set_Writer()
            self.Is_Data_good()
            self.Set_Layout()
            self.Save_to_excel()
            print('------------------Execution is complete---------------------')
        except:
            self.status = 'Failure'
            raise

    def make_copy(self):
        self.sumfile = 'Swiggy_summary.xlsx'
        shutil.copy('Swiggy Summary-output1.xlsx',self.sumfile)

    def Do_Accounting(self):

        all_files = glob.glob(self.path1 + "\*.xlsx")
        fi = []
        self.book = openpyxl.load_workbook(self.sumfile)

        for filename in all_files:
            #--------Read Input files--------------
            print (filename)
            self.week = pd.read_excel(filename, skiprows =8)

            #--------------Restaurant header details-------------------
            self.head = pd.read_excel(filename, nrows=4,usecols=[1,2,3,4])

            sh = self.book['Sheet1']
            sh['C2'] =self.head.iloc[1,1]
            sh['C3'] =self.head.iloc[0,3]
            sh['C4'] =''
            sh['C5'] =self.head.iloc[0,1]
            #------------------------------------------------------------

            #-------Week days list---------------------------------------
            startday = (self.week['Order Date']).str.split('-',expand=True)[2].str.split(' ',expand = True)[0].min()
            endday  = (self.week['Order Date']).str.split('-',expand=True)[2].str.split(' ',expand = True)[0].max()
            year = (self.week['Order Date']).str.split('-',expand=True)[0][1]
            month = (self.week['Order Date']).str.split('-',expand=True)[1][1]
            print (startday +'-'+month+'-'+year +' to ' + endday+'-'+month+'-'+year)

            col = int(int(startday)/7) + 3
            #Period
            sh.cell(row=6, column=col).value = (startday)+ '-'+ month +'-'+ year+' to ' + endday + '-'+ month +'-'+ year

            #Bill Value
            ctcs = list(self.week.filter(regex='CTCS').columns)[0]
            bv = list(self.week.filter(regex='Bill value').columns)[0]
            sh.cell(row=8, column=col).value = self.week[bv].sum() - self.week[self.week[ctcs]==0][bv].sum()
            #Discount
            disc = list(self.week.filter(regex='Restaurant Trade Discount').columns)[0]
            sh.cell(row=9, column=col).value = self.week[disc].sum() - self.week[self.week[ctcs]==0][disc].sum()
            #Net Bill Value
            #GST on Sale
            gst = list(self.week.filter(regex='GST liability').columns)[0]
            sh.cell(row=13, column=col).value = self.week[gst].sum()
            #Service fee as per Invoice
            Sfee = list(self.week.filter(regex='Swiggy platform service').columns)[0]
            sh.cell(row=16, column=col).value = self.week[Sfee].sum()
            #Lead Generation Fee
            lgfee = list(self.week.filter(regex='Lead Generation').columns)[0]
            sh.cell(row=17, column=col).value = self.week[lgfee].sum()
            #IVRS Cost
            ivrs = list(self.week.filter(regex='Call Center Service Fees').columns)[0]
            sh.cell(row=18, column=col).value = self.week[ivrs].sum()
            #Effective platform fee
            try:
                epsf = list(self.week.filter(regex='Effective platform support fee').columns)[0]
                sh.cell(row=19, column=col).value = self.week[epsf].sum()
            except:
                sh.cell(row=19, column=col).value = 0
            #Total Service Fee before gst
            tsfgst = list(self.week.filter(regex='Total Swiggy Service').columns)[0]
            sh.cell(row=20, column=col).value = self.week[tsfgst].sum()
            #CGST
            cgst = list(self.week.filter(regex='CGST').columns)[0]
            sh.cell(row=21, column=col).value = self.week[cgst].sum()
            #SGST
            sgst = list(self.week.filter(regex='SGST').columns)[0]
            sh.cell(row=22, column=col).value = self.week[sgst].sum()
            #Total service fee
            tsf = list(self.week.filter(regex='Total Swiggy service fee').columns)[0]
            sh.cell(row=23, column=col).value = self.week[tsf].sum()
            #Refund to customer
            ref = list(self.week.filter(regex='Refund To Customer').columns)[0]
            sh.cell(row=24, column=col).value = self.week[ref].sum()
            #CTCS
            ctcs = list(self.week.filter(regex='CTCS').columns)[0]
            sh.cell(row=25, column=col).value = self.week[ctcs].sum()
            #STCS
            stcs = list(self.week.filter(regex='STCS').columns)[0]
            sh.cell(row=26, column=col).value = self.week[stcs].sum()
            #Amount recieved aginst cancelled orders
            can_ord= list(self.week.filter(regex='before TCS deduction').columns)[0]
            sh.cell(row=27, column=col).value = self.week[can_ord].sum() - self.week[self.week[ctcs]==0][can_ord].sum()
            #Amount to be credited as per input files
            amount= list(self.week.filter(regex='after TCS deduction').columns)[0]
            sh.cell(row=29, column=col).value = self.week[amount].sum()

        all_files = glob.glob(self.path1 + "\*.csv")
        for filename in all_files:
            self.dump =  pd.read_csv(filename,skiprows=5,usecols=range(0,30))
            self.dump = self.dump[self.dump['Order-status']=='delivered']

            #-------days----------2019-11-04 00:19:56----------30-11-2019 22:31
            startday = (self.dump['Order-delivery-time']).str.split('-',expand=True)[0].str.split(' ',expand = True)[0].min()
            endday  = (self.dump['Order-delivery-time']).str.split('-',expand=True)[0].str.split(' ',expand = True)[0].max()
            year = (self.dump['Order-delivery-time']).str.split('-',expand=True)[2][1].split(' ')[0]
            month = (self.dump['Order-delivery-time']).str.split('-',expand=True)[1][1].split(' ')[0]
            print (startday +'-'+month+'-'+year +' to ' + endday+'-'+month+'-'+year)

            col = int(int(startday)/7) + 3
            #Period
            sh.cell(row=6, column=col).value = (startday)+ '-'+ month +'-'+ year+' to ' + endday + '-'+ month +'-'+ year

            #Bill Value
            bv = list(self.dump.filter(regex='Total-bill-amount').columns)[0]
            sh.cell(row=8, column=col).value = self.dump[bv].sum()
            #Discount
            disc = list(self.week.filter(regex='Restaurant Trade Discount').columns)[0]
            sh.cell(row=9, column=col).value = self.week[disc].sum() - self.week[self.week[ctcs]==0][disc].sum()
            #Net Bill Value
            #GST on Sale
            gst = list(self.week.filter(regex='GST liability').columns)[0]
            sh.cell(row=13, column=col).value = self.week[gst].sum()
            #Service fee as per Invoice
            Sfee = list(self.week.filter(regex='Swiggy platform service').columns)[0]
            sh.cell(row=16, column=col).value = self.week[Sfee].sum()
            #Lead Generation Fee
            lgfee = list(self.week.filter(regex='Lead Generation').columns)[0]
            sh.cell(row=17, column=col).value = self.week[lgfee].sum()
            #IVRS Cost
            ivrs = list(self.week.filter(regex='Call Center Service Fees').columns)[0]
            sh.cell(row=18, column=col).value = self.week[ivrs].sum()
            #Effective platform fee
            try:
                epsf = list(self.week.filter(regex='Effective platform support fee').columns)[0]
                sh.cell(row=19, column=col).value = self.week[epsf].sum()
            except:
                sh.cell(row=19, column=col).value = 0
            #Total Service Fee before gst
            tsfgst = list(self.week.filter(regex='Total Swiggy Service').columns)[0]
            sh.cell(row=20, column=col).value = self.week[tsfgst].sum()
            #CGST
            cgst = list(self.week.filter(regex='CGST').columns)[0]
            sh.cell(row=21, column=col).value = self.week[cgst].sum()
            #SGST
            sgst = list(self.week.filter(regex='SGST').columns)[0]
            sh.cell(row=22, column=col).value = self.week[sgst].sum()
            #Total service fee
            tsf = list(self.week.filter(regex='Total Swiggy service fee').columns)[0]
            sh.cell(row=23, column=col).value = self.week[tsf].sum()
            #Refund to customer
            ref = list(self.week.filter(regex='Refund To Customer').columns)[0]
            sh.cell(row=24, column=col).value = self.week[ref].sum()
            #CTCS
            ctcs = list(self.week.filter(regex='CTCS').columns)[0]
            sh.cell(row=25, column=col).value = self.week[ctcs].sum()
            #STCS
            stcs = list(self.week.filter(regex='STCS').columns)[0]
            sh.cell(row=26, column=col).value = self.week[stcs].sum()
            #Amount recieved aginst cancelled orders
            can_ord= list(self.week.filter(regex='before TCS deduction').columns)[0]
            sh.cell(row=27, column=col).value = self.week[can_ord].sum() - self.week[self.week[ctcs]==0][can_ord].sum()
            #Amount to be credited as per input files
            amount= list(self.week.filter(regex='after TCS deduction').columns)[0]
            sh.cell(row=29, column=col).value = self.week[amount].sum()



    def Set_Writer(self):
        self.writer = pd.ExcelWriter('SDLR.xlsx', engine='xlsxwriter')


    def Is_Data_good(self):
        pass

    def Set_Layout(self):
        pass

    def Save_to_excel(self):
        self.writer.save()
        self.book.save(self.sumfile)

#-------------Report 6 --Consumption Analysis--------------------------
class Consumption_Analysis(RepGen):

    # Initialize the class
    def __init__(self, os, ps, cs, months, years, file_suffix):

        self.osf = os
        self.psf = ps
        self.csf = cs
        self.months = months
        self.years = years
        self.df = []
        self.monthnames = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        self.status = 'Failure'
        try:

            rd = pd.read_excel(cs['cur'], nrows=3)
            self.j3 = rd.iloc[1, 9]
            self.j4 = rd.iloc[2, 9]
            self.fname = 'media/uploads/' + file_suffix + 'Consumption_Analysis_Report.xlsx'
            self.writer = pd.ExcelWriter(self.fname, engine='xlsxwriter')

            # create table dictionary
            self.table = {}

            # create stack dataframe
            self.stack = pd.DataFrame()
            i = 0
            self.header = pd.read_excel(self.psf, nrows=4, header=None)
            header = self.header[[0]]
            client_name = self.header[0][0]

            # month_yr = self.header[0][2]#.str.split('-')
            # m = re.search('- (.+?),', month_yr)
            # y = month_yr[-4:]
            # if m:
            #     found = m.group(1)
            mn_yr = self.monthnames[int(months[3]) - 1] + ' ' + years[3]
            rep_name = "Consumption Analysis"
            head_json = {"customer": client_name, "report_type": rep_name, "month_year": mn_yr}

            for os, cs in list(zip(self.osf.values(), self.csf.values())):
                # Read purchase files from QB
                self.qb = pd.read_excel(self.psf, skiprows=4)
                # drop irrelevant column
                self.qb.drop(columns=self.qb.columns[0], inplace=True)
                self.qb = self.qb[~self.qb.isnull().all(axis=1)]
                # Delete unnecessary rows
                self.qb = self.qb[~(self.qb.isnull().sum(axis=1) >= 9)]

                # reduce to current data

                self.qb = self.qb[self.qb.Date.str.split('/').str[1] == self.months[i]]

                # Read closing stock files
                try:
                    self.cs = pd.read_excel(cs)
                    for name in self.cs.columns:
                        if not name == 'ItemName':
                            self.cs.rename(columns={name: 'cs_' + name}, inplace=True)
                except:
                    self.cs = pd.DataFrame(columns=["ItemName"])
                # Read Opening stock files
                try:
                    self.os = pd.read_excel(os)
                    for name in self.os.columns:
                        if not name == 'ItemName':
                            self.os.rename(columns={name: 'os_' + name}, inplace=True)
                except:
                    self.os = pd.DataFrame(columns=["ItemName"])
                # create intermediate consumption Dataframe
                self.cons = pd.DataFrame(columns=['ItemName',
                                                  'UOM',
                                                  'Date',
                                                  'Transaction_type',
                                                  'BillNo',
                                                  'Supplier',
                                                  'Qty',
                                                  'Rate',
                                                  'Amount',
                                                  'FileTag'])
                ########## Consumption tables [1] #############
                self.Do_Accounting(cs, i)

                i += 1

            ########## Consumption change by Quantity [2] #############
            lst = []

            print("testt")
            print(self.table)

            for key in self.table:
                try:
                    lst.append(pd.DataFrame(self.table[key])[['ItemName', 'consumption in quantity']])
                except:
                    pass
            print(lst)
            try:
                self.df = pd.concat(lst[0:-1])
            except Exception as e:
                print(e)
                pass
            print("test self df")
            print(self.df)

            self.change = pd.merge(self.df.groupby('ItemName').mean(), lst[-1], on='ItemName',
                                   how='inner')
            print("test self df end ")
            self.change.rename(columns={'consumption in quantity_x': 'Average consumption in last 3 months (Qtn)',
                                        'consumption in quantity_y': 'consumptionn this month (Qtn)'},
                               inplace=True)
            self.change['% change in consumption'] = 100 * (self.change['consumptionn this month (Qtn)'] -
                                                            self.change[
                                                                'Average consumption in last 3 months (Qtn)']) / (
                                                         self.change['Average consumption in last 3 months (Qtn)'])
            self.change = round(self.change, 2)
            self.change = self.change.fillna(0)

            ########## Consumption change by Value [3] ###############
            lst = []

            for key in self.table:

                try:
                    lst.append(pd.DataFrame(self.table[key])[['ItemName', 'consumption in value']])
                except:
                    pass
            try:
                self.df = pd.concat(lst[0:-1])
            except:
                pass
            self.rchange = pd.merge(self.df.groupby('ItemName').mean(), lst[-1], on='ItemName',
                                    how='inner')
            self.rchange.rename(columns={'consumption in value_x': 'Average consumption in last 3 months (in value)',
                                         'consumption in value_y': 'consumption this month (in value)'},
                                inplace=True)
            self.rchange['% change in consumption'] = 100 * (self.rchange['consumption this month (in value)'] -
                                                             self.rchange[
                                                                 'Average consumption in last 3 months (in value)']) / (
                                                          self.rchange[
                                                              'Average consumption in last 3 months (in value)'])
            self.rchange = round(self.rchange, 2)
            self.rchange = self.rchange.fillna(0)

            ########## Change in Consumption by % of sales [4] ###############
            lst = []
            for key in self.table:
                try:
                    lst.append(pd.DataFrame(self.table[key])[
                                   ['ItemName', 'consumption in value', '% of consumption to total sales']])
                except:
                    pass
            try:
                self.df = pd.concat(lst[0:-1])
            except:
                pass
            self.schange = pd.merge(self.df.groupby('ItemName').sum(), lst[-1], on='ItemName',
                                    how='inner')
            self.schange.rename(columns={'consumption in value_x': 'Average consumption in last 3 months (% of Sales)',
                                         '% of consumption to total sales_y': 'consumption this month (% of Sales)'},
                                inplace=True)
            self.schange['Average consumption in last 3 months (% of Sales)'] = 100 * self.schange[
                'Average consumption in last 3 months (% of Sales)'] / self.j4

            self.schange['% change in consumption'] = 100 * (self.schange['consumption this month (% of Sales)'] -
                                                             self.schange[
                                                                 'Average consumption in last 3 months (% of Sales)']) / (
                                                          self.schange[
                                                              'Average consumption in last 3 months (% of Sales)'])
            self.schange.drop(columns=['% of consumption to total sales_x', 'consumption in value_y'], inplace=True)
            self.schange = round(self.schange, 2)
            self.schange = self.schange.fillna(0)
            ########## Change in Prices Currrent Vs Last 3 months ###############

            # weighted avg price for last 3 months
            self.pchange = pd.DataFrame(columns=['ItemName',
                                                 'Average price in the last 3 months',
                                                 'Price this month',
                                                 '% change in prices'])

            df1 = self.stack[self.stack.Date.str.split('/').str[1] != self.months[3]]

            df2 = self.stack[self.stack.Date.str.split('/').str[1] == self.months[3]]

            self.pchange = pd.merge(df1.groupby('ItemName', as_index=False).agg({'Qty': 'sum', 'Amount': 'sum'}),
                                    df2.groupby('ItemName', as_index=False).agg({'Qty': 'sum', 'Amount': 'sum'}),
                                    how='inner',
                                    on='ItemName')

            self.pchange['Average price in the last 3 months'] = self.pchange['Amount_x'] / self.pchange['Qty_x']

            self.pchange['Price this month'] = self.pchange['Amount_y'] / self.pchange['Qty_y']

            self.pchange.drop(columns=['Amount_x', 'Amount_y', 'Qty_x', 'Qty_y'], inplace=True)

            self.pchange['% change in prices'] = 100 * (
                    self.pchange['Price this month'] - self.pchange['Average price in the last 3 months']) / \
                                                 self.pchange['Average price in the last 3 months']

            self.pchange = round(self.pchange, 2)
            self.pchange = self.pchange.fillna(0)

            ########## Current month price varation  ###############

            self.mchange = pd.merge(
                df2.groupby('ItemName', as_index=False).Rate.min().rename(columns={'Rate': 'Least price this month'}),
                df2.groupby('ItemName', as_index=False).Rate.max().rename(columns={'Rate': 'Highest price this month'}),
                on='ItemName',
                how='inner')
            df = self.stack[self.stack.Date.str.split('/').str[1] == self.months[3]]
            df = df[['ItemName', 'Rate', 'Qty', 'Amount']]
            df = df.groupby('ItemName', as_index=False).agg({'Qty': 'sum', 'Amount': 'sum', 'Rate': 'std'})
            df['weighted rate'] = df['Amount'] / df['Qty']
            self.mchange['Price fluctuation index'] = 100 * df['Rate'] / df['weighted rate']
            self.mchange = round(self.mchange, 2)
            self.mchange = self.mchange.fillna(0)

            #         print(self.monthnames[int(self.months[3])]+str(self.years[3])[-2:])
            ##################################################################################
            self.table['Change_in_Consumption_By_Quantity'] = self.change.to_dict(orient="records")
            self.table['Change_in_Consumption_By_Quantity'].insert(0, {'Name': 'Change in Consumption (By Quantity)'})

            self.table['Change_in_Consumption_By_Value'] = self.rchange.to_dict(orient="records")
            self.table['Change_in_Consumption_By_Value'].insert(0, {'Name': 'Change in Consumption (By Value)'})

            self.table['Change_in_Consumption_By_per_of_sales'] = self.schange.to_dict(orient="records")
            self.table['Change_in_Consumption_By_per_of_sales'].insert(0, {
                'Name': 'Change in Consumption (By % of sales)'})

            self.table['Change_in_Prices_compared_to_previous_months'] = self.pchange.to_dict(orient="records")
            self.table['Change_in_Prices_compared_to_previous_months'].insert(0, {
                'Name': 'Change in Prices compared to previous 3 months'})

            self.table['Change_in_Prices_Comparision_within_the_month'] = self.mchange.to_dict(orient="records")
            self.table['Change_in_Prices_Comparision_within_the_month'].insert(0, {
                'Name': 'Change in Prices Comparision within the month'})
            ##### Convert tables to Json format ############
            ### code written by dvg (next 36 lines)
            ###removing unwanted consumption tables (previous 3 months) from output
            table_d1 = collections.OrderedDict(self.table)
            tableitems = list(table_d1.items())
            tableitems[0:3] = []
            keys1 = ['ItemName', 'consumption in value', 'UOM']
            keys2 = ['prices purchased for', 'vendors purchased from', 'consumption in quantity',
                     '% of consumption to total sales']
            keys3 = ['ItemName', '% change in consumption']
            keys4 = ['Average consumption in last 3 months (Qtn)', 'consumptionn this month (Qtn)']
            keys5 = ['Average consumption in last 3 months (in value)', 'consumption this month (in value)']
            keys6 = ['Average consumption in last 3 months (% of Sales)', 'consumption this month (% of Sales)']
            keys7 = ['ItemName', '% change in prices']
            keys8 = ['ItemName', 'Price fluctuation index']
            keys9 = ['Average price in the last 3 months', 'Price this month']
            keys10 = ['Least price this month', 'Highest price this month']
            list3 = []
            list4 = []
            list5 = []
            list6 = []
            list7 = []
            list8 = []
            for obj in tableitems[0][1]:
                list3.append({'ItemSummary': {x: obj[x] for x in keys1}, 'detailedList': {y: obj[y] for y in keys2}})
            for j in list3:
                j['ItemSummary']['consumption'] = j['ItemSummary']['consumption in value']
                j['ItemSummary'].pop('consumption in value')
                j['detailedList'] = [{'Name': k, 'value': v} for k, v in list(j['detailedList'].items())]
            for obj in tableitems[1][1][1:]:
                list4.append({'ItemSummary': {x: obj[x] for x in keys3}, 'detailedList': {y: obj[y] for y in keys4}})
            for j in list4:
                j['ItemSummary']['consumptionVariation'] = j['ItemSummary']['% change in consumption']
                j['ItemSummary'].pop('% change in consumption')
                j['detailedList'] = [{'Name': k, 'value': v} for k, v in list(j['detailedList'].items())]
            for obj in tableitems[2][1][1:]:
                list5.append({'ItemSummary': {x: obj[x] for x in keys3}, 'detailedList': {y: obj[y] for y in keys5}})
            for j in list5:
                j['ItemSummary']['consumptionVariation'] = j['ItemSummary']['% change in consumption']
                j['ItemSummary'].pop('% change in consumption')
                j['detailedList'] = [{'Name': k, 'value': v} for k, v in list(j['detailedList'].items())]
            for obj in tableitems[3][1][1:]:
                list6.append({'ItemSummary': {x: obj[x] for x in keys3}, 'detailedList': {y: obj[y] for y in keys6}})
            for j in list6:
                j['ItemSummary']['consumptionVariation'] = j['ItemSummary']['% change in consumption']
                j['ItemSummary'].pop('% change in consumption')
                j['detailedList'] = [{'Name': k, 'value': v} for k, v in list(j['detailedList'].items())]
            for obj in tableitems[4][1][1:]:
                list7.append({'ItemSummary': {x: obj[x] for x in keys7}, 'detailedList': {y: obj[y] for y in keys9}})
            for j in list7:
                j['ItemSummary']['PriceVariation'] = j['ItemSummary']['% change in prices']
                j['ItemSummary'].pop('% change in prices')
                j['detailedList'] = [{'Name': k, 'value': v} for k, v in list(j['detailedList'].items())]
            for obj in tableitems[5][1][1:]:
                list8.append({'ItemSummary': {x: obj[x] for x in keys8}, 'detailedList': {y: obj[y] for y in keys10}})
            for j in list8:
                j['ItemSummary']['PriceFluctuation'] = j['ItemSummary']['Price fluctuation index']
                j['ItemSummary'].pop('Price fluctuation index')
                j['detailedList'] = [{'Name': k, 'value': v} for k, v in list(j['detailedList'].items())]

            finaldict = {
                "header": head_json,
                "consumptionTable": list3,
                "Change_in_Consumption_By_Quantity": list4,
                "Change_in_Consumption_By_Value": list5,
                "Change_in_Consumption_By_per_of_sales": list6,
                "Change_in_Prices_compared_to_previous_months": list7,
                "Change_in_Prices_Comparision_within_the_month": list8
            }
            self.tab1 = pd.DataFrame(tableitems[0][1])
            self.tab1.to_excel(self.writer, index=False, sheet_name='Consumption Table', startrow=4)
            self.tab2 = pd.DataFrame(self.table['Change_in_Consumption_By_Quantity'][1:])
            self.tab2.to_excel(self.writer, index=False, sheet_name='Consumption Summary', startrow=4)
            self.tab3 = pd.DataFrame(self.table['Change_in_Consumption_By_Value'][1:])
            self.tab3.to_excel(self.writer, index=False, sheet_name='Consumption Summary', startrow=14)
            self.tab4 = pd.DataFrame(self.table['Change_in_Consumption_By_per_of_sales'][1:])
            self.tab4.to_excel(self.writer, index=False, sheet_name='Consumption Summary', startrow=24)
            self.tab5 = pd.DataFrame(self.table['Change_in_Prices_compared_to_previous_months'][1:])
            self.tab5.to_excel(self.writer, index=False, sheet_name='Consumption Summary', startrow=34)
            self.tab6 = pd.DataFrame(self.table['Change_in_Prices_Comparision_within_the_month'][1:])
            self.tab6.to_excel(self.writer, index=False, sheet_name='Consumption Summary', startrow=44)
            self.Set_Writer()
            self.Set_Layout()
            self.Save_to_excel()
            # header = pd.read_excel(file_name,nrows=4,header=None)
            # header = header[[0]]
            # client_name = header[0][0]
            # month_yr = header[0][2]
            # m = re.search('- (.+?),', month_yr)
            # y = month_yr[-4:]
            # if m:
            #     found = m.group(1)
            # mn_yr = found[:3] + ' ' +y
            # rep_name = 'Consumption Analysis'
            # head_json = {'customer':client_name, 'report_type':rep_name, 'month-year':mn_yr }
            self.json = simplejson.dumps(finaldict,ignore_nan=True)
            self.status = 'Success'
            print(self.json)
        except:
            self.status = 'Failure'
            raise

    def Set_Writer(self):

        self.workbook = self.writer.book
        self.worksheet = self.writer.sheets['Consumption Table']
        self.page2 = self.writer.sheets['Consumption Summary']

    def Do_Accounting(self, cs, ind):

        # Assign values to consumption file from purchases file
        self.cons['ItemName'] = self.qb['Product/Service']
        self.cons['Transaction_type'] = self.qb['Transaction Type']
        self.cons['Date'] = self.qb['Date']
        self.cons['BillNo'] = self.qb['No.']
        self.cons['Supplier'] = self.qb['Supplier']
        self.cons['Qty'] = self.qb['Qty']
        self.cons['Rate'] = self.qb['Rate']
        self.cons['Amount'] = self.qb['Amount']
        # File tag for purchases
        self.cons['FileTag'] = 'Purchases'

        '''
        Consumption Analysis :
            Consumption = Opening Stock + Purchases - Closing stock 
        '''
        if self.cons.empty:
            self.cons = pd.DataFrame(columns=["ItemName"])
        dfs = [self.os, self.cons, self.cs]
        self.mr = reduce(lambda left, right: pd.merge(left, right, on=['ItemName'],
                                                      how='inner'), dfs)
        self.stack = self.stack.append(self.mr, ignore_index=True)
        lst = set(self.cs['ItemName']) & set(self.cons['ItemName']) & set(self.cs['ItemName'])
        self.result = self.consumption_table(lst, self.mr)
        print("self result")
        print(self.result)

        #         print(ind)

        # store list of dictionaries in a table dictionary
        self.table[
            'consumptionTable-' + self.monthnames[int(self.months[ind]) - 1] + str(self.years[ind])[-2:]] = self.result

    def consumption_table(self, lst, dfinal):

        # initialize lines of rows to be written in table
        lines = []

        # loop on common list of items from closing stock and purchase items to generate consumption table
        for item in lst:
            hf = dfinal[dfinal['ItemName'] == item].reset_index()
            row = {}
            row['ItemName'] = item

            row['UOM'] = list(hf.os_UOM.unique())
            row['prices purchased for'] = sorted(list(hf.Rate.dropna().unique()))
            row['vendors purchased from'] = list(hf.Supplier.dropna().unique())
            try:
                row['consumption in quantity'] = hf['cs_Stock in Store'][0] + hf.Qty.sum() - \
                                                 hf['cs_Stock in Store'][0]
            except:
                row['consumption in quantity'] = ' '
            row['consumption in value'] = hf['os_Total'][0] + hf.Amount.sum() - \
                                          hf['cs_Total'][0]

            row['% of consumption to total sales'] = round(100 * row['consumption in value'] / self.j3, 2)
            lines.append(row)

        return (lines)

    def Set_Layout(self):
        self.monthnames = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        mn_yr = self.monthnames[int(self.months[3]) - 1] + ' ' + self.years[3]
        print("mn_yr")
        print(mn_yr)
        merge_cells(self.workbook, self.worksheet, 'A1', 'G1', str(self.header[0][0]))
        merge_cells(self.workbook, self.worksheet, 'A2', 'G2', 'Consumption Details')
        merge_cells(self.workbook, self.worksheet, 'A3', 'G3', str(mn_yr))

        merge_cells(self.workbook, self.page2, 'A1', 'D1', str(self.header[0][0]))
        merge_cells(self.workbook, self.page2, 'A2', 'D2', 'Consumption Summary')
        merge_cells(self.workbook, self.page2, 'A3', 'D3', str(mn_yr))

    def Save_to_excel(self):
        self.writer.save()

class Sales_Recon(RepGen):

    def __init__(self, bankTranx_file, swiggy_folder, zomato_folder, dunzo_folder, nearbuy_folder, dineout_folder, swiggy_pos, zomato_pos, dunzo_pos, nearbuy_pos, dineout_pos,file_suffix=None):
        super().__init__(bankTranx_file)
        try:
            self.rootdir=os.path.dirname(os.path.abspath(bankTranx_file))

            self.swiggy_reco(bankTranx_file, swiggy_folder)

            self.zomato_reco(bankTranx_file, zomato_folder)

            self.dunzo_reco(bankTranx_file, dunzo_folder)

            self.nearbuy_reco(bankTranx_file, nearbuy_folder)

            self.dineout_reco(bankTranx_file, dineout_folder)

            self.table = {}
            self.setupjson(swiggy_pos, zomato_pos, dunzo_pos, nearbuy_pos, dineout_pos)

            self.status = 'Success'
            print('------------------Execution is complete---------------------')
        except:
            self.status = 'Failure'
            raise
        #####------created reco summary-----###########

    ###functions for individual recos  ##############
    def swiggy_reco(self, bankTranx_file, swiggy_folder):
        self.df_bank = pd.read_excel(bankTranx_file)
        #         with ZipFile (swiggy_file, 'r') as zip:
        listOfiles = os.listdir(swiggy_folder)
        nOffiles = len(listOfiles)
        #---------#create template columns list and empty output dataframe----------------##--------------------------#
        ColList_template = ['Order Date', 'Order No', 'Order Status', 'Bill value\n(without taxes & discount)', 'Merchant Trade Discount',
                            'Merchant coupon Discount share', 'Net bill value (without taxes)', 'Customer payable\n(Net bill value after taxes & discount)',
                            'Swiggy platform service fee', 'Lead Generation fee @ 7.5%', 'Call Center Service Fees', 'Delivery Service fee\n(sponsored by merchant)',
                            'Effective delivery service fee\n(net of commission & tax)', 'Total Swiggy service fee\n(including taxes)', 'Refund To Customer',
                            'Net Payable Amount (before TCS and TDS deduction)', 'CTCS', 'STCS', 'ITCS', 'TDS',
                            'Net Payable Amount (after TCS and TDS deduction)', 'Order Type']
        self.df_swgy1 = pd.DataFrame(np.array([['Sales (Exclusive of GST) (Exclusive of cancelled and failed orders)', ""], ['less: Discounts', ""], \
                                               ['Sales (Exclusive of GST) after discounts', ""], ['add: GST on sales (@5%)', ""], ['Sales (Inclusive of GST)', ""], \
                                               ['', ""],['Commission and other charges (Inclusive of GST)', ""], ['Swiggy fee (Commission)', ""], ['IVRS cost', ""], \
                                               ['Platform fee', ""], ['Lead generation fee', ""],['Carousel and advertising', ""], ['Total Commission and other Charges', ""], \
                                               ['', ""], ['Cancelled order costs', ""], ['add: Money received against cancelled orders', ""], ['less: Refund to customer', ""], \
                                               ['Total Cancelled order costs', ""],['', ""], ['A. Tax adjustments', ""], ['less: TCS Deducted', ""], ['B. Other adjustments', ""], \
                                               ["Opening week's adjustments", ""], ["Closing week's adjustments", ""], ["Total of Adjustments", ""], ["", ""], \
                                               ["Expected receipt in Bank Statement", ""], ["Actual Receipt in Bank Statement", ""], ["Difference", ""]]), \
                                     columns=['Particulars', 'percentage'])
        # #Step3------------
        # #count no. of input files and their dateranges
        filesinfo = []
        cnt = 1
        for item in listOfiles:
            item=swiggy_folder+"/"+item
            self.df1 = pd.read_excel(item, nrows=6,header=None)
            strDate = self.df1.loc[3,2]
            endDate = self.df1.loc[4,2]
            sDate = strDate[0:5]
            eDate = endDate[0:5]
            # #Step4----------------
            # #read input data and add new columns with calculated values in reconc dataframe
            self.df2 = pd.read_excel(item, skiprows=8)
            self.df2.fillna("")
            cnt = cnt+1
            colname = str(sDate) + ' to ' + str(eDate)
            for cname in ColList_template:
                stringMatch = cname
                self.df2.columns = [stringMatch if col.startswith(stringMatch) else col for col in self.df2.columns]
            for colm in self.df2.columns:
                self.df2[colm] = self.df2[colm].replace(" ","")
            self.df2 = self.df2.filter(ColList_template)
            self.df_swgy1['newcol'] = np.nan
            try:
                self.df_swgy1.loc[0,'newcol'] = self.df2['Bill value\n(without taxes & discount)'].sum()- self.df2[self.df2['CTCS'] == 0]['Bill value\n(without taxes & discount)'].sum()
            except:
                None
            #                 self.df_swgy1.loc[0,'newcol'] = 0
            try:
                self.df_swgy1.loc[1,'newcol'] = -(self.df2['Merchant coupon Discount share'].sum()- self.df2[self.df2['CTCS'] == 0]['Merchant coupon Discount share'].sum())
            except:
                self.df_swgy1.loc[1,'newcol'] = 0
            self.df_swgy1.loc[2,'newcol'] = self.df_swgy1.loc[0,'newcol'] + self.df_swgy1.loc[1,'newcol']
            self.df_swgy1.loc[3,'newcol'] = 0.05 * self.df_swgy1.loc[2,'newcol']
            self.df_swgy1.loc[4,'newcol'] = self.df_swgy1.loc[2,'newcol'] + self.df_swgy1.loc[3,'newcol']
            self.df_swgy1.loc[6,'newcol'] = ""
            try:
                self.df_swgy1.loc[7,'newcol'] = 1.18 * self.df2['Swiggy platform service fee'].sum()
            except:
                self.df_swgy1.loc[7,'newcol'] = 0
            try:
                self.df_swgy1.loc[8,'newcol'] = 1.18 * self.df2['Call Center Service Fees'].sum()
            except:
                self.df_swgy1.loc[8,'newcol'] = 0
            try:
                self.df_swgy1.loc[9,'newcol'] = 1.18 * self.df2['Effective delivery service fee(net of commission & tax)'].sum()
            except:
                self.df_swgy1.loc[9,'newcol'] = 0
            try:
                self.df_swgy1.loc[10,'newcol'] = 1.18 * self.df2['Lead Generation fee @ 7.5%'].sum()
            except:
                self.df_swgy1.loc[10,'newcol'] = 0
            self.df_swgy1.loc[11,'newcol'] = ""
            self.df_swgy1.loc[12,'newcol'] = self.df_swgy1.loc[7,'newcol'] + self.df_swgy1.loc[8,'newcol'] + self.df_swgy1.loc[9,'newcol'] + self.df_swgy1.loc[10,'newcol']
            self.df_swgy1.loc[14,'newcol'] = ""
            try:
                self.df_swgy1.loc[15,'newcol'] = self.df2[self.df2['CTCS'] == 0]['Net Payable Amount (after TCS deduction)'].sum()
            except:
                self.df_swgy1.loc[15,'newcol'] = 0
            try:
                self.df_swgy1.loc[16,'newcol'] = -self.df2['Refund To Custome'].sum()
            except:
                self.df_swgy1.loc[16,'newcol'] = 0
            self.df_swgy1.loc[17,'newcol'] = self.df_swgy1.loc[15,'newcol'] + self.df_swgy1.loc[16,'newcol']
            self.df_swgy1.loc[19,'newcol'] = ""
            try:
                self.df_swgy1.loc[20,'newcol'] = -(self.df2['CTCS'].sum() + self.df2['STCS'].sum())
            except:
                self.df_swgy1.loc[20,'newcol'] = 0
            self.df_swgy1.loc[21,'newcol'] = ""
            self.df_swgy1.loc[5,'newcol'] = ""
            self.df_swgy1.loc[13,'newcol'] = ""
            self.df_swgy1.loc[18,'newcol'] = ""
            self.df_swgy1.loc[25,'newcol'] = ""
            self.df_swgy1.loc[22,'newcol'] = 0
            self.df_swgy1.loc[23,'newcol'] = 0
            self.df_swgy1.loc[24,'newcol'] = self.df_swgy1.loc[20,'newcol'] + self.df_swgy1.loc[22,'newcol'] + self.df_swgy1.loc[23,'newcol']
            self.df_swgy1.loc[26,'newcol'] = self.df_swgy1.loc[4,'newcol'] - self.df_swgy1.loc[12,'newcol'] + self.df_swgy1.loc[17,'newcol'] + self.df_swgy1.loc[24,'newcol']
            self.df_swgy1.loc[27,'newcol'] = self.df_bank[(self.df_bank['PlatformName'] == 'Swiggy') & (self.df_bank['InvoiceFile_tag'] == (cnt-1))]['Credit/Deposit'].sum()
            self.df_swgy1.loc[28,'newcol'] = self.df_swgy1.loc[27,'newcol'] - self.df_swgy1.loc[26,'newcol']
            self.df_swgy1 = self.df_swgy1.rename(columns = {'newcol': colname})

        percentList = [4,7,8,9,10,12,16,17,20,22,23,24,26,27,28]
        rwlist = [0,1,2,3,7,8,9,10,15,16,20,22,23,27]
        blankpercentlist = [0,1,2,3,6,11,14,15,19,21]
        nanlist = [6, 11, 14, 19, 21, 5, 13, 18, 25]
        self.df_swgy1['total'] = np.nan
        for k in rwlist:
            self.df_swgy1.loc[k,'total'] = self.df_swgy1.iloc[k, -(nOffiles+1):-1].sum()
        for i in nanlist:
            self.df_swgy1.loc[i,'total'] = ""
        self.df_swgy1.loc[4,'total'] = self.df_swgy1.loc[2,'total'] + self.df_swgy1.loc[3,'total']
        self.df_swgy1.loc[12,'total'] = self.df_swgy1.loc[7,'total'] + self.df_swgy1.loc[8,'total'] + self.df_swgy1.loc[9,'total'] + self.df_swgy1.loc[10,'total']
        self.df_swgy1.loc[17,'total'] = self.df_swgy1.loc[15,'total'] + self.df_swgy1.loc[16,'total']
        self.df_swgy1.loc[24,'total'] = self.df_swgy1.loc[20,'total'] + self.df_swgy1.loc[22,'total'] + self.df_swgy1.loc[23,'total']
        self.df_swgy1.loc[26,'total'] = self.df_swgy1.loc[4,'total'] - self.df_swgy1.loc[12,'total'] + self.df_swgy1.loc[17,'total'] + self.df_swgy1.loc[24,'total']
        self.df_swgy1.loc[28,'total'] = self.df_swgy1.loc[27,'total'] - self.df_swgy1.loc[26,'total']
        for i in percentList:
            self.df_swgy1.loc[i,'percentage'] = (self.df_swgy1.loc[i,'total']/self.df_swgy1.loc[4,'total']) * 100
        for j in blankpercentlist:
            self.df_swgy1.loc[j,'percentage'] = ""
        self.df_swgy1 = self.df_swgy1[[c for c in self.df_swgy1 if c not in ['percentage']] + ['percentage']]
    #         display(self.df_swgy1)

    def zomato_reco(self, bankTranx_file, zomato_folder):
        df_bank = pd.read_excel(bankTranx_file)
        # read zipfile and extract contents
        #         with ZipFile (zomato_file, 'r') as zip:
        listOfiles = os.listdir(zomato_folder)
        nOffiles = len(listOfiles)
        #step2#----------
        #---------#create template columns list and empty output dataframe----------------##--------------------------#
        ColList_template = ['Order ID', 'Order Date', 'Order status', 'Gross revenue (INR)', 'Customer discount amount (INR)',
                            'Commissionable amount (INR)', 'Commission %', 'Commission value (INR)', 'Convenience fee (INR)',
                            'Cancellation Refund (INR)', 'Taxes on Zomato fees (INR)', 'Tax collected at source (INR)']
        self.df_zom1 = pd.DataFrame(np.array([['Sales (Exclusive of GST) (Exclusive of cancelled and failed orders)', ""], ['less: Discounts', ""], \
                                              ['Sales (Exclusive of GST) after discounts', ""], ['add: GST on sales (@5%)', ""], ['Sales (Inclusive of GST)', ""], \
                                              ['', ""],['Commission and other charges (Inclusive of GST)', ""], ['Zomato fee (Commission)', ""], ['Convenience fee (Payment gateway charges)', ""], \
                                              ['Advertising and other Marketing Charges', ""], ['Total Commission and other Charges', ""], \
                                              ['', ""], ['Cancelled order costs', ""], ['add: Money received against cancelled orders', ""], ['less: Refund to customer', ""], \
                                              ['Total Cancelled order costs', ""],['', ""], ['A. Tax adjustments', ""], ['less: TCS Deducted', ""], ['B. Other adjustments', ""], \
                                              ["Opening week's adjustments", ""], ["Closing week's adjustments", ""], ["Total of Adjustments", ""], ["", ""], \
                                              ["Expected receipt in Bank Statement", ""], ["Actual Receipt in Bank Statement", ""], ["Difference", ""]]), \
                                    columns=['Particulars', 'percentage'])
        # #Step3------------
        # #count no. of input files and their dateranges
        filesinfo = []
        cnt = 1
        for item in listOfiles:
            item=zomato_folder+"/"+item
            self.df1 = pd.read_excel(item)
            strDate = self.df1['Order Date'].min()
            endDate = self.df1['Order Date'].max()
            sDate = str(strDate.date())[8:] + '-' + str(strDate.date())[5:7]
            eDate = str(endDate.date())[8:] + '-' + str(endDate.date())[5:7]
            # #Step4----------------
            # #read input data and add new columns with calculated values in reconc dataframe
            cnt = cnt+1
            colname = str(sDate) + ' to ' + str(eDate)
            for cname in ColList_template:
                stringMatch = cname
                self.df1.columns = [stringMatch if col.startswith(stringMatch) else col for col in self.df1.columns]
            for colm in self.df1.columns:
                self.df1[colm] = self.df1[colm].replace(" ","")
            self.df1 = self.df1.filter(ColList_template)
            self.df_zom1['newcol'] = np.nan
            try:
                self.df_zom1.loc[0,'newcol'] = 100/105 * self.df1['Gross revenue (INR)'].sum()
            except:
                self.df_zom1.loc[0,'newcol'] = 0
            try:
                self.df_zom1.loc[1,'newcol'] = -(100/105 * self.df1['Customer discount amount (INR)'].sum())
            except:
                self.df_zom1.loc[1,'newcol'] = 0
            self.df_zom1.loc[2,'newcol'] = self.df_zom1.loc[0,'newcol'] + self.df_zom1.loc[1,'newcol']
            self.df_zom1.loc[3,'newcol'] = 0.05 * self.df_zom1.loc[2,'newcol']
            self.df_zom1.loc[4,'newcol'] = self.df_zom1.loc[2,'newcol'] + self.df_zom1.loc[3,'newcol']
            try:
                self.df_zom1.loc[7,'newcol'] = 1.18 * self.df1['Commission value (INR)'].sum()
            except:
                self.df_zom1.loc[7,'newcol'] = 0
            try:
                self.df_zom1.loc[8,'newcol'] = 1.18 * self.df1['Convenience fee (INR)'].sum()
            except:
                self.df_zom1.loc[8,'newcol'] = 0
            self.df_zom1.loc[9,'newcol'] = 0
            self.df_zom1.loc[10,'newcol'] = self.df_zom1.loc[7,'newcol'] + self.df_zom1.loc[8,'newcol'] + self.df_zom1.loc[9,'newcol']
            try:
                self.df_zom1.loc[13,'newcol'] = self.df1['Cancellation refund (INR)'].sum()
            except:
                self.df_zom1.loc[13,'newcol'] = 0
            self.df_zom1.loc[14,'newcol'] = 0
            self.df_zom1.loc[15,'newcol'] = self.df_zom1.loc[13,'newcol'] + self.df_zom1.loc[14,'newcol']
            try:
                self.df_zom1.loc[18,'newcol'] = -(self.df1['Tax collected at source (INR)'].sum())
            except:
                self.df_zom1.loc[18,'newcol'] = 0
            self.df_zom1.loc[6,'newcol'] = ""
            self.df_zom1.loc[12,'newcol'] = ""
            self.df_zom1.loc[17,'newcol'] = ""
            self.df_zom1.loc[19,'newcol'] = ""
            self.df_zom1.loc[5,'newcol'] = ""
            self.df_zom1.loc[11,'newcol'] = ""
            self.df_zom1.loc[16,'newcol'] = ""
            self.df_zom1.loc[23,'newcol'] = ""
            self.df_zom1.loc[20,'newcol'] = 0
            self.df_zom1.loc[21,'newcol'] = 0
            self.df_zom1.loc[22,'newcol'] = self.df_zom1.loc[18,'newcol'] + self.df_zom1.loc[20,'newcol'] + self.df_zom1.loc[21,'newcol']
            self.df_zom1.loc[24,'newcol'] = self.df_zom1.loc[4,'newcol'] - self.df_zom1.loc[10,'newcol'] + self.df_zom1.loc[15,'newcol'] + self.df_zom1.loc[22,'newcol']
            self.df_zom1.loc[25,'newcol'] = self.df_bank[(self.df_bank['PlatformName'] == 'Zomato') & (self.df_bank['InvoiceFile_tag'] == (cnt-1))]['Credit/Deposit'].sum()
            self.df_zom1.loc[26,'newcol'] = self.df_zom1.loc[25,'newcol'] - self.df_zom1.loc[24,'newcol']
            self.df_zom1 = self.df_zom1.rename(columns = {'newcol': colname})
        percentList = [4,7,8,9,10,13,14,15,18,20,21,22,24,25,26]
        rwlist = [0,1,2,3,7,8,9,13,14,18,20,21,25]
        blankpercentlist = [0,1,2,3,6,12,17,19]
        nanlist = [6, 12, 17, 19, 5, 11, 16, 23]
        self.df_zom1['total'] = np.nan
        for k in rwlist:
            self.df_zom1.loc[k,'total'] = self.df_zom1.iloc[k, -(nOffiles+1):-1].sum()
        for i in nanlist:
            self.df_zom1.loc[i,'total'] = ""
        self.df_zom1.loc[4,'total'] = self.df_zom1.loc[2,'total'] + self.df_zom1.loc[3,'total']
        self.df_zom1.loc[10,'total'] = self.df_zom1.loc[7,'total'] + self.df_zom1.loc[8,'total'] + self.df_zom1.loc[9,'total']
        self.df_zom1.loc[15,'total'] = self.df_zom1.loc[13,'total'] + self.df_zom1.loc[14,'total']
        self.df_zom1.loc[22,'total'] = self.df_zom1.loc[18,'total'] + self.df_zom1.loc[20,'total'] + self.df_zom1.loc[21,'total']
        self.df_zom1.loc[24,'total'] = self.df_zom1.loc[4,'total'] - self.df_zom1.loc[10,'total'] + self.df_zom1.loc[15,'total'] + self.df_zom1.loc[22,'total']
        self.df_zom1.loc[26,'total'] = self.df_zom1.loc[25,'total'] - self.df_zom1.loc[24,'total']
        for i in percentList:
            self.df_zom1.loc[i,'percentage'] = (self.df_zom1.loc[i,'total']/self.df_zom1.loc[4,'total']) * 100
        for j in blankpercentlist:
            self.df_zom1.loc[j,'percentage'] = ""
        self.df_zom1 = self.df_zom1[[c for c in self.df_zom1 if c not in ['percentage']] + ['percentage']]
    #         display(self.df_zom1)

    def dunzo_reco(self, bankTranx_file, dunzo_folder):
        df_bank = pd.read_excel(bankTranx_file)
        # read zipfile and extract contents
        #         with ZipFile (dunzo_file, 'r') as zip:
        listOfiles = os.listdir(dunzo_folder)
        nOffiles = len(listOfiles)
        #step2#----------
        #---------#create template columns list and empty output dataframe----------------##--------------------------#
        ColList_template = ['Merchant Id', ' Order Date', ' Order Id', ' Txn Id', ' Bill No.', ' Order Status', ' Tax', ' Product Total Amount',
                            'Product Discount', ' Product Commission']
        self.df_Dunz1 = pd.DataFrame(np.array([['Sales (Exclusive of GST) (Exclusive of cancelled and failed orders)', ""], ['less: Discounts', ""], \
                                               ['Sales (Exclusive of GST) after discounts', ""], ['add: GST on sales (@5%)', ""], ['Sales (Inclusive of GST)', ""], \
                                               ['', ""],['Commission and other charges (Inclusive of GST)', ""], ['Commission', ""], \
                                               ['Advertising and other Marketing Charges', ""], ['Total Commission and other Charges', ""], \
                                               ['', ""], ['Cancelled order costs', ""], ['add: Money received against cancelled orders', ""], ['less: Refund to customer', ""], \
                                               ['Total Cancelled order costs', ""],['', ""], ['A. Tax adjustments', ""], ['less: TCS Deducted', ""], ['B. Other adjustments', ""], \
                                               ["Opening week's adjustments", ""], ["Closing week's adjustments", ""], ["Total of Adjustments", ""], ["", ""], \
                                               ["Expected receipt in Bank Statement", ""], ["Actual Receipt in Bank Statement", ""], ["Difference", ""]]), \
                                     columns=['Particulars', 'percentage'])
        # #Step3------------
        # #count no. of input files and their dateranges
        filesinfo = []
        cnt = 1
        for item in listOfiles:
            item=dunzo_folder+"/"+item
            self.df1 = pd.read_excel(item)
            strDate = self.df1['Order Date'].min()
            endDate = self.df1['Order Date'].max()
            format_str = '%a %b %d %H:%M:%S %Y' # The format
            strDate = datetime.strptime(strDate, format_str)
            endDate = datetime.strptime(endDate, format_str)
            sDate = str(strDate.date())[8:] + '-' + str(strDate.date())[5:7]
            eDate = str(endDate.date())[8:] + '-' + str(endDate.date())[5:7]
            # #Step4----------------
            # #read input data and add new columns with calculated values in reconc dataframe
            cnt = cnt+1
            colname = str(sDate) + ' to ' + str(eDate)
            for cname in ColList_template:
                stringMatch = cname
                self.df1.columns = [stringMatch if col.startswith(stringMatch) else col for col in self.df1.columns]
            for colm in self.df1.columns:
                self.df1[colm] = self.df1[colm].replace(" ","")
            #     self.df1 = self.df1.filter(ColList_template)
            self.df_Dunz1['newcol'] = np.nan
            try:
                self.df_Dunz1.loc[0,'newcol'] = 100/105 * self.df1['Product Total Amount'].sum()
            except:
                self.df_Dunz1.loc[0,'newcol'] = 0
            try:
                self.df_Dunz1.loc[1,'newcol'] = -(100/105 * self.df1['Product Discount'].sum())
            except:
                self.df_Dunz1.loc[1,'newcol'] = 0
            self.df_Dunz1.loc[2,'newcol'] = self.df_Dunz1.loc[0,'newcol'] + self.df_Dunz1.loc[1,'newcol']
            self.df_Dunz1.loc[3,'newcol'] = 0.05 * self.df_Dunz1.loc[2,'newcol']
            self.df_Dunz1.loc[4,'newcol'] = self.df_Dunz1.loc[2,'newcol'] + self.df_Dunz1.loc[3,'newcol']
            try:
                self.df_Dunz1.loc[7,'newcol'] = 1.18 * self.df1['Product Commission'].sum()
            except:
                self.df_Dunz1.loc[7,'newcol'] = 0
            self.df_Dunz1.loc[8,'newcol'] = 0
            self.df_Dunz1.loc[9,'newcol'] = self.df_Dunz1.loc[7,'newcol'] + self.df_Dunz1.loc[8,'newcol']
            self.df_Dunz1.loc[12,'newcol'] = 0
            self.df_Dunz1.loc[13,'newcol'] = 0
            self.df_Dunz1.loc[14,'newcol'] = self.df_Dunz1.loc[12,'newcol'] + self.df_Dunz1.loc[13,'newcol']
            try:
                self.df_Dunz1.loc[17,'newcol'] = -(self.df1['Tax'].sum())
            except:
                self.df_Dunz1.loc[17,'newcol'] = 0
            self.df_Dunz1.loc[6,'newcol'] = ""
            self.df_Dunz1.loc[11,'newcol'] = ""
            self.df_Dunz1.loc[16,'newcol'] = ""
            self.df_Dunz1.loc[18,'newcol'] = ""
            self.df_Dunz1.loc[5,'newcol'] = ""
            self.df_Dunz1.loc[10,'newcol'] = ""
            self.df_Dunz1.loc[15,'newcol'] = ""
            self.df_Dunz1.loc[22,'newcol'] = ""
            self.df_Dunz1.loc[19,'newcol'] = 0
            self.df_Dunz1.loc[20,'newcol'] = 0
            self.df_Dunz1.loc[21,'newcol'] = self.df_Dunz1.loc[17,'newcol'] + self.df_Dunz1.loc[19,'newcol'] + self.df_Dunz1.loc[20,'newcol']
            self.df_Dunz1.loc[23,'newcol'] = self.df_Dunz1.loc[4,'newcol'] - self.df_Dunz1.loc[9,'newcol'] + self.df_Dunz1.loc[14,'newcol'] + self.df_Dunz1.loc[21,'newcol']
            self.df_Dunz1.loc[24,'newcol'] = self.df_bank[(self.df_bank['PlatformName'] == 'Dunzo') & (self.df_bank['InvoiceFile_tag'] == (cnt-1))]['Credit/Deposit'].sum()
            self.df_Dunz1.loc[25,'newcol'] = self.df_Dunz1.loc[24,'newcol'] - self.df_Dunz1.loc[23,'newcol']
            self.df_Dunz1 = self.df_Dunz1.rename(columns = {'newcol': colname})

        percentList = [4,7,8,9,12,13,14,17,19,20,21,23,24,25]
        rwlist = [0,1,2,3,7,8,12,13,17,19,20,23,24]
        blankpercentlist = [0,1,2,3,6,11,16,18]
        nanlist = [6, 11, 16, 18, 5, 10, 15, 22]
        self.df_Dunz1['total'] = np.nan
        for k in rwlist:
            self.df_Dunz1.loc[k,'total'] = self.df_Dunz1.iloc[k, -(nOffiles+1):-1].sum()
        for i in nanlist:
            self.df_Dunz1.loc[i,'total'] = ""
        self.df_Dunz1.loc[4,'total'] = self.df_Dunz1.loc[2,'total'] + self.df_Dunz1.loc[3,'total']
        self.df_Dunz1.loc[9,'total'] = self.df_Dunz1.loc[7,'total'] + self.df_Dunz1.loc[8,'total']
        self.df_Dunz1.loc[14,'total'] = self.df_Dunz1.loc[12,'total'] + self.df_Dunz1.loc[13,'total']
        self.df_Dunz1.loc[21,'total'] = self.df_Dunz1.loc[17,'total'] + self.df_Dunz1.loc[19,'total'] + self.df_Dunz1.loc[20,'total']
        self.df_Dunz1.loc[23,'total'] = self.df_Dunz1.loc[4,'total'] - self.df_Dunz1.loc[9,'total'] + self.df_Dunz1.loc[14,'total'] + self.df_Dunz1.loc[21,'total']
        self.df_Dunz1.loc[25,'total'] = self.df_Dunz1.loc[24,'total'] - self.df_Dunz1.loc[23,'total']
        for i in percentList:
            self.df_Dunz1.loc[i,'percentage'] = (self.df_Dunz1.loc[i,'total']/self.df_Dunz1.loc[4,'total']) * 100
        for j in blankpercentlist:
            self.df_Dunz1.loc[j,'percentage'] = ""
        self.df_Dunz1 = self.df_Dunz1[[c for c in self.df_Dunz1 if c not in ['percentage']] + ['percentage']]
    #         display(self.df_Dunz1)

    def nearbuy_reco(self, bankTranx_file, nearbuy_folder):
        df_bank = pd.read_excel(bankTranx_file)
        # read zipfile and extract contents
        #         with ZipFile (nearbuy_file, 'r') as zip:
        listOfiles = os.listdir(nearbuy_folder)
        nOffiles = len(listOfiles)
        #step2#----------
        #---------#create template columns list and empty output dataframe----------------##--------------------------#
        ColList_template = ['Voucher Code', 'Order ID', 'Business Account ID', 'Customer Price (Rs)', 'Taxable value for TCS', 'TCS deducted',
                            'Net Payable to Merchant', 'Discount by Merchant', 'Redemption Date', 'Payment initiated On', 'Paid Date', 'Transaction ID']
        self.df_Near1 = pd.DataFrame(np.array([['Customer price Sales (Exclusive of GST) (Exclusive of cancelled and failed orders)', ""], ['less: Discounts', ""], \
                                               ['Sales (Exclusive of GST) after discounts', ""], ['add: GST on sales (@5%)', ""], ['Sales (Inclusive of GST)', ""], \
                                               ['', ""],['Commission and other charges (Inclusive of GST)', ""], ['Commission @11.80% (Inclusive of GST)', ""], \
                                               ['Advertising and other Marketing Charges', ""], ['Total Commission and other Charges', ""], \
                                               ['', ""], ['Cancelled order costs', ""], ['add: Money received against cancelled orders', ""], ['less: Refund to customer', ""], \
                                               ['Total Cancelled order costs', ""],['', ""], ['A. Tax adjustments', ""], ['less: TCS Deducted', ""], ['B. Other adjustments', ""], \
                                               ["Opening week's adjustments", ""], ["Closing week's adjustments", ""], ["Total of Adjustments", ""], ["", ""], \
                                               ["Expected receipt in Bank Statement", ""], ["Actual Receipt in Bank Statement", ""], ["Difference", ""]]), \
                                     columns=['Particulars', 'percentage'])
        # #Step3------------
        # #count no. of input files and their dateranges
        filesinfo = []
        cnt = 1
        for item in listOfiles:
            item=nearbuy_folder+"/"+item
            self.df1 = pd.read_excel(item)
            strDate = self.df1['Paid Date'].min()
            endDate = self.df1['Paid Date'].max()
            sDate = str(strDate.date())[8:] + '-' + str(strDate.date())[5:7]
            eDate = str(endDate.date())[8:] + '-' + str(endDate.date())[5:7]
            # #Step4----------------
            # #read input data and add new columns with calculated values in reconc dataframe
            cnt = cnt+1
            colname = str(sDate) + ' to ' + str(eDate)
            for cname in ColList_template:
                stringMatch = cname
                self.df1.columns = [stringMatch if col.startswith(stringMatch) else col for col in self.df1.columns]
            for colm in self.df1.columns:
                self.df1[colm] = self.df1[colm].replace(" ","")
            #     self.df1 = self.df1.filter(ColList_template)
            self.df_Near1['newcol'] = np.nan
            try:
                self.df_Near1.loc[0,'newcol'] = 100/105 * self.df1['Customer Price (Rs)'].sum()
            except:
                self.df_Near1.loc[0,'newcol'] = 0
            try:
                self.df_Near1.loc[1,'newcol'] = -(100/105 * self.df1['Discount by Merchant'].sum())
            except:
                self.df_Near1.loc[1,'newcol'] = 0
            self.df_Near1.loc[2,'newcol'] = self.df_Near1.loc[0,'newcol'] + self.df_Near1.loc[1,'newcol']
            self.df_Near1.loc[3,'newcol'] = 0.05 * self.df_Near1.loc[2,'newcol']
            self.df_Near1.loc[4,'newcol'] = self.df_Near1.loc[2,'newcol'] + self.df_Near1.loc[3,'newcol']
            self.df_Near1.loc[7,'newcol'] = 0.118 * self.df_Near1.loc[4,'newcol']
            self.df_Near1.loc[8,'newcol'] = 0
            self.df_Near1.loc[9,'newcol'] = self.df_Near1.loc[7,'newcol'] + self.df_Near1.loc[8,'newcol']
            self.df_Near1.loc[12,'newcol'] = 0
            self.df_Near1.loc[13,'newcol'] = 0
            self.df_Near1.loc[14,'newcol'] = self.df_Near1.loc[12,'newcol'] + self.df_Near1.loc[13,'newcol']
            try:
                self.df_Near1.loc[17,'newcol'] = -(self.df1['TCS deducted'].sum())
            except:
                self.df_Near1.loc[17,'newcol'] = 0
            self.df_Near1.loc[6,'newcol'] = ""
            self.df_Near1.loc[11,'newcol'] = ""
            self.df_Near1.loc[16,'newcol'] = ""
            self.df_Near1.loc[18,'newcol'] = ""
            self.df_Near1.loc[5,'newcol'] = ""
            self.df_Near1.loc[10,'newcol'] = ""
            self.df_Near1.loc[15,'newcol'] = ""
            self.df_Near1.loc[22,'newcol'] = ""
            self.df_Near1.loc[19,'newcol'] = 0
            self.df_Near1.loc[20,'newcol'] = 0
            self.df_Near1.loc[21,'newcol'] = self.df_Near1.loc[17,'newcol'] + self.df_Near1.loc[19,'newcol'] + self.df_Near1.loc[20,'newcol']
            self.df_Near1.loc[23,'newcol'] = self.df_Near1.loc[4,'newcol'] - self.df_Near1.loc[9,'newcol'] + self.df_Near1.loc[14,'newcol'] + self.df_Near1.loc[21,'newcol']
            self.df_Near1.loc[24,'newcol'] = self.df_bank[(self.df_bank['PlatformName'] == 'Nearbuy') & (self.df_bank['InvoiceFile_tag'] == (cnt-1))]['Credit/Deposit'].sum()
            self.df_Near1.loc[25,'newcol'] = self.df_Near1.loc[24,'newcol'] - self.df_Near1.loc[23,'newcol']
            self.df_Near1 = self.df_Near1.rename(columns = {'newcol': colname})

        percentList = [4,7,8,9,12,13,14,17,19,20,21,23,24,25]
        rwlist = [0,1,2,3,7,8,12,13,17,19,20,23,24]
        blankpercentlist = [0,1,2,3,6,11,16,18]
        nanlist = [6, 11, 16, 18, 5, 10, 15, 22]
        self.df_Near1['total'] = np.nan
        for k in rwlist:
            self.df_Near1.loc[k,'total'] = self.df_Near1.iloc[k, -(nOffiles+1):-1].sum()
        for i in nanlist:
            self.df_Near1.loc[i,'total'] = ""
        self.df_Near1.loc[4,'total'] = self.df_Near1.loc[2,'total'] + self.df_Near1.loc[3,'total']
        self.df_Near1.loc[9,'total'] = self.df_Near1.loc[7,'total'] + self.df_Near1.loc[8,'total']
        self.df_Near1.loc[14,'total'] = self.df_Near1.loc[12,'total'] + self.df_Near1.loc[13,'total']
        self.df_Near1.loc[21,'total'] = self.df_Near1.loc[17,'total'] + self.df_Near1.loc[19,'total'] + self.df_Near1.loc[20,'total']
        self.df_Near1.loc[23,'total'] = self.df_Near1.loc[4,'total'] - self.df_Near1.loc[9,'total'] + self.df_Near1.loc[14,'total'] + self.df_Near1.loc[21,'total']
        self.df_Near1.loc[25,'total'] = self.df_Near1.loc[24,'total'] - self.df_Near1.loc[23,'total']
        for i in percentList:
            self.df_Near1.loc[i,'percentage'] = (self.df_Near1.loc[i,'total']/self.df_Near1.loc[4,'total']) * 100
        for j in blankpercentlist:
            self.df_Near1.loc[j,'percentage'] = ""
        self.df_Near1 = self.df_Near1[[c for c in self.df_Near1 if c not in ['percentage']] + ['percentage']]
    #         display(self.df_Near1)

    def dineout_reco(self, bankTranx_file, dineout_folder):
        df_bank = pd.read_excel(bankTranx_file)
        # read zipfile and extract contents
        #         with ZipFile (dineout_file, 'r') as zip:
        #         listOfiles = os.listdir("5-Dineout Invoice/"), swiggy_folder, zomato_folder, dunzo_folder, nearbuy_folder, dineout_folder
        listOfiles = os.listdir(dineout_folder)
        nOffiles = len(listOfiles)
        #step2#----------
        #---------#create template columns list and empty output dataframe----------------##--------------------------#
        ColList_template = ['Order ID', 'Order Date', 'Order status', 'Gross revenue (INR)', 'Customer discount amount (INR)',
                            'Commissionable amount (INR)', 'Commission %', 'Commission value (INR)', 'Convenience fee (INR)',
                            'Cancellation Refund (INR)', 'Taxes on Platform fees (INR)', 'Tax collected at source (INR)']
        self.df_Dnot1 = pd.DataFrame(np.array([['Customer price Sales (Exclusive of GST) (Exclusive of cancelled and failed orders)', ""], ['less: Discounts', ""], \
                                               ['Sales (Exclusive of GST) after discounts', ""], ['add: GST on sales (@5%)', ""], ['Sales (Inclusive of GST)', ""], \
                                               ['', ""],['Commission and other charges (Inclusive of GST)', ""], ['Commission (Inclusive of GST)', ""], \
                                               ['Advertising and other Marketing Charges', ""], ['Total Commission and other Charges', ""], \
                                               ['', ""], ['Cancelled order costs', ""], ['add: Money received against cancelled orders', ""], ['less: Refund to customer', ""], \
                                               ['Total Cancelled order costs', ""],['', ""], ['A. Tax adjustments', ""], ['less: TCS Deducted', ""], ['less: TDS Deducted', ""], ['B. Other adjustments', ""], \
                                               ["Opening week's adjustments", ""], ["Closing week's adjustments", ""], ["Total of Adjustments", ""], ["", ""], \
                                               ["Expected receipt in Bank Statement", ""], ["Actual Receipt in Bank Statement", ""], ["Difference", ""]]), \
                                     columns=['Particulars', 'percentage'])
        # #Step3------------
        # #count no. of input files and their dateranges
        filesinfo = []
        cnt = 1
        for item in listOfiles:
            item=dineout_folder+"/"+item
            self.df1 = pd.read_excel(item)
            strDate = self.df1['Order Date'].min()
            endDate = self.df1['Order Date'].max()
            sDate = str(strDate.date())[8:] + '-' + str(strDate.date())[5:7]
            eDate = str(endDate.date())[8:] + '-' + str(endDate.date())[5:7]
            # #Step4----------------
            # #read input data and add new columns with calculated values in reconc dataframe
            cnt = cnt+1
            colname = str(sDate) + ' to ' + str(eDate)
            for cname in ColList_template:
                stringMatch = cname
                self.df1.columns = [stringMatch if col.startswith(stringMatch) else col for col in self.df1.columns]
            for colm in self.df1.columns:
                self.df1[colm] = self.df1[colm].replace(" ","")
            self.df1 = self.df1.filter(ColList_template)
            self.df_Dnot1['newcol'] = np.nan
            try:
                self.df_Dnot1.loc[0,'newcol'] = 100/105 * self.df1['Gross revenue (INR)'].sum()
            except:
                self.df_Dnot1.loc[0,'newcol'] = 0
            try:
                self.df_Dnot1.loc[1,'newcol'] = -(100/105 * self.df1['Customer discount amount (INR)'].sum())
            except:
                self.df_Dnot1.loc[1,'newcol'] = 0
            self.df_Dnot1.loc[2,'newcol'] = self.df_Dnot1.loc[0,'newcol'] + self.df_Dnot1.loc[1,'newcol']
            self.df_Dnot1.loc[3,'newcol'] = 0.05 * self.df_Dnot1.loc[2,'newcol']
            self.df_Dnot1.loc[4,'newcol'] = self.df_Dnot1.loc[2,'newcol'] + self.df_Dnot1.loc[3,'newcol']
            try:
                self.df_Dnot1.loc[7,'newcol'] = 1.18 * self.df1['Commission value (INR)'].sum()
            except:
                self.df_Dnot1.loc[7,'newcol'] = 0
            try:
                self.df_Dnot1.loc[8,'newcol'] = 1.18 * self.df1['Convenience fee (INR)'].sum()
            except:
                self.df_Dnot1.loc[8,'newcol'] = 0
            self.df_Dnot1.loc[9,'newcol'] = self.df_Dnot1.loc[7,'newcol'] + self.df_Dnot1.loc[8,'newcol']
            try:
                self.df_Dnot1.loc[12,'newcol'] = self.df1['Cancellation refund (INR)'].sum()
            except:
                self.df_Dnot1.loc[12,'newcol'] = 0
            self.df_Dnot1.loc[13,'newcol'] = 0
            self.df_Dnot1.loc[14,'newcol'] = self.df_Dnot1.loc[12,'newcol'] + self.df_Dnot1.loc[13,'newcol']
            try:
                self.df_Dnot1.loc[17,'newcol'] = -(self.df1['less: TCS Deducted'].sum())
            except:
                self.df_Dnot1.loc[17,'newcol'] = 0
            try:
                self.df_Dnot1.loc[18,'newcol'] = -(self.df1['less: TDS Deducted'].sum())
            except:
                self.df_Dnot1.loc[18,'newcol'] = 0
            self.df_Dnot1.loc[6,'newcol'] = ""
            self.df_Dnot1.loc[11,'newcol'] = ""
            self.df_Dnot1.loc[16,'newcol'] = ""
            self.df_Dnot1.loc[19,'newcol'] = ""
            self.df_Dnot1.loc[5,'newcol'] = ""
            self.df_Dnot1.loc[10,'newcol'] = ""
            self.df_Dnot1.loc[15,'newcol'] = ""
            self.df_Dnot1.loc[23,'newcol'] = ""
            self.df_Dnot1.loc[20,'newcol'] = 0
            self.df_Dnot1.loc[21,'newcol'] = 0
            self.df_Dnot1.loc[22,'newcol'] = self.df_Dnot1.loc[17,'newcol'] + self.df_Dnot1.loc[18,'newcol'] + self.df_Dnot1.loc[20,'newcol'] + self.df_Dnot1.loc[21,'newcol']
            self.df_Dnot1.loc[24,'newcol'] = self.df_Dnot1.loc[4,'newcol'] - self.df_Dnot1.loc[9,'newcol'] + self.df_Dnot1.loc[14,'newcol'] + self.df_Dnot1.loc[22,'newcol']
            self.df_Dnot1.loc[25,'newcol'] = self.df_bank[(self.df_bank['PlatformName'] == 'Dineout') & (self.df_bank['InvoiceFile_tag'] == (cnt-1))]['Credit/Deposit'].sum()
            self.df_Dnot1.loc[26,'newcol'] = self.df_Dnot1.loc[25,'newcol'] - self.df_Dnot1.loc[24,'newcol']
            self.df_Dnot1 = self.df_Dnot1.rename(columns = {'newcol': colname})
        percentList = [4,7,8,9,12,13,14,17,18,20,21,22,24,25,26]
        rwlist = [0,1,2,3,7,8,12,13,17,18,20,21,25]
        blankpercentlist = [0,1,2,3,6,11,16,19]
        nanlist = [6, 11, 16, 19, 5, 10, 15, 23]
        self.df_Dnot1['total'] = np.nan
        for k in rwlist:
            self.df_Dnot1.loc[k,'total'] = self.df_Dnot1.iloc[k, -(nOffiles+1):-1].sum()
        for i in nanlist:
            self.df_Dnot1.loc[i,'total'] = ""
        self.df_Dnot1.loc[4,'total'] = self.df_Dnot1.loc[2,'total'] + self.df_Dnot1.loc[3,'total']
        self.df_Dnot1.loc[9,'total'] = self.df_Dnot1.loc[7,'total'] + self.df_Dnot1.loc[8,'total']
        self.df_Dnot1.loc[14,'total'] = self.df_Dnot1.loc[12,'total'] + self.df_Dnot1.loc[13,'total']
        self.df_Dnot1.loc[22,'total'] = self.df_Dnot1.loc[17,'total'] + self.df_Dnot1.loc[18,'total'] + self.df_Dnot1.loc[20,'total'] + self.df_Dnot1.loc[21,'total']
        self.df_Dnot1.loc[24,'total'] = self.df_Dnot1.loc[4,'total'] - self.df_Dnot1.loc[9,'total'] + self.df_Dnot1.loc[14,'total'] + self.df_Dnot1.loc[22,'total']
        self.df_Dnot1.loc[26,'total'] = self.df_Dnot1.loc[25,'total'] - self.df_Dnot1.loc[24,'total']
        for i in percentList:
            self.df_Dnot1.loc[i,'percentage'] = (self.df_Dnot1.loc[i,'total']/self.df_Dnot1.loc[4,'total']) * 100
        for j in blankpercentlist:
            self.df_Dnot1.loc[j,'percentage'] = ""
        self.df_Dnot1 = self.df_Dnot1[[c for c in self.df_Dnot1 if c not in ['percentage']] + ['percentage']]


    def setupjson(self,swiggy_pos, zomato_pos, dunzo_pos, nearbuy_pos, dineout_pos):
        self.df_recoSum = pd.DataFrame(np.array([['Sales (Inclusive of GST) as per channel',"","","","","","","","","","","",""],
                                                 ['Sales (inclucive of GST) as per POS',"","","","","","","","","","","",""],
                                                 ['Net Sales Difference',"","","","","","","","","","","",""], ['',"","","","","","","","","","","",""],
                                                 ['Commission & other charges',"","","","","","","","","","","",""], ['',"","","","","","","","","","","",""],
                                                 ['Cancelled orders cost',"","","","","","","","","","","",""], ['',"","","","","","","","","","","",""],
                                                 ['Other adjustments',"","","","","","","","","","","",""], ['',"","","","","","","","","","","",""],
                                                 ['Expected net receipts',"","","","","","","","","","","",""],
                                                 ['Actual receipts',"","","","","","","","","","","",""],
                                                 ['Difference',"","","","","","","","","","","",""]]), \
                                       columns=['Particulars', 'Swiggy', '% to total Swiggy sales', 'Zomato', '% to total Zomato sales', 'Dunzo',
                                                '% to total Dunzo sales', 'Nearbuy', '% to total Nearbuy sales', 'Dineout', '% to total Dineout sales',
                                                'Total', '% to Total sales of all channels'])
        #swiggy summary
        self.df_recoSum.loc[0, 'Swiggy'] = self.df_swgy1.loc[4,'total']
        self.df_recoSum.loc[1, 'Swiggy'] = int(swiggy_pos)
        self.df_recoSum.loc[2, 'Swiggy'] = self.df_recoSum.loc[0, 'Swiggy'] - self.df_recoSum.loc[1, 'Swiggy']
        self.df_recoSum.loc[4, 'Swiggy'] = self.df_swgy1.loc[12,'total']
        self.df_recoSum.loc[6, 'Swiggy'] = self.df_swgy1.loc[17,'total']
        self.df_recoSum.loc[8, 'Swiggy'] = self.df_swgy1.loc[24,'total']
        self.df_recoSum.loc[10, 'Swiggy'] = self.df_swgy1.loc[26,'total']
        self.df_recoSum.loc[11, 'Swiggy'] = self.df_swgy1.loc[27,'total']
        self.df_recoSum.loc[12, 'Swiggy'] = self.df_recoSum.loc[10, 'Swiggy'] - self.df_recoSum.loc[11, 'Swiggy']
        #Zomato summary
        self.df_recoSum.loc[0, 'Zomato'] = self.df_zom1.loc[4,'total']
        self.df_recoSum.loc[1, 'Zomato'] = int(zomato_pos)
        self.df_recoSum.loc[2, 'Zomato'] = self.df_recoSum.loc[0, 'Zomato'] - self.df_recoSum.loc[1, 'Zomato']
        self.df_recoSum.loc[4, 'Zomato'] = self.df_zom1.loc[10,'total']
        self.df_recoSum.loc[6, 'Zomato'] = self.df_zom1.loc[15,'total']
        self.df_recoSum.loc[8, 'Zomato'] = self.df_zom1.loc[22,'total']
        self.df_recoSum.loc[10, 'Zomato'] = self.df_zom1.loc[24,'total']
        self.df_recoSum.loc[11, 'Zomato'] = self.df_zom1.loc[25,'total']
        self.df_recoSum.loc[12, 'Zomato'] = self.df_recoSum.loc[10, 'Zomato'] - self.df_recoSum.loc[11, 'Zomato']
        #Dunzo summary
        self.df_recoSum.loc[0, 'Dunzo'] = self.df_Dunz1.loc[4,'total']
        self.df_recoSum.loc[1, 'Dunzo'] = int(dunzo_pos)
        self.df_recoSum.loc[2, 'Dunzo'] = self.df_recoSum.loc[0, 'Dunzo'] - self.df_recoSum.loc[1, 'Dunzo']
        self.df_recoSum.loc[4, 'Dunzo'] = self.df_Dunz1.loc[9,'total']
        self.df_recoSum.loc[6, 'Dunzo'] = self.df_Dunz1.loc[14,'total']
        self.df_recoSum.loc[8, 'Dunzo'] = self.df_Dunz1.loc[21,'total']
        self.df_recoSum.loc[10, 'Dunzo'] = self.df_Dunz1.loc[23,'total']
        self.df_recoSum.loc[11, 'Dunzo'] = self.df_Dunz1.loc[24,'total']
        self.df_recoSum.loc[12, 'Dunzo'] = self.df_recoSum.loc[10, 'Dunzo'] - self.df_recoSum.loc[11, 'Dunzo']
        #Nearbuy summary
        self.df_recoSum.loc[0, 'Nearbuy'] = self.df_Near1.loc[4,'total']
        self.df_recoSum.loc[1, 'Nearbuy'] = int(nearbuy_pos)
        self.df_recoSum.loc[2, 'Nearbuy'] = self.df_recoSum.loc[0, 'Nearbuy'] - self.df_recoSum.loc[1, 'Nearbuy']
        self.df_recoSum.loc[4, 'Nearbuy'] = self.df_Near1.loc[9,'total']
        self.df_recoSum.loc[6, 'Nearbuy'] = self.df_Near1.loc[14,'total']
        self.df_recoSum.loc[8, 'Nearbuy'] = self.df_Near1.loc[21,'total']
        self.df_recoSum.loc[10, 'Nearbuy'] = self.df_Near1.loc[23,'total']
        self.df_recoSum.loc[11, 'Nearbuy'] = self.df_Near1.loc[24,'total']
        self.df_recoSum.loc[12, 'Nearbuy'] = self.df_recoSum.loc[10, 'Nearbuy'] - self.df_recoSum.loc[11, 'Nearbuy']
        #Dineout summary
        self.df_recoSum.loc[0, 'Dineout'] = self.df_Dnot1.loc[4,'total']
        self.df_recoSum.loc[1, 'Dineout'] = int(dineout_pos)
        self.df_recoSum.loc[2, 'Dineout'] = self.df_recoSum.loc[0, 'Dineout'] - self.df_recoSum.loc[1, 'Dineout']
        self.df_recoSum.loc[4, 'Dineout'] = self.df_Dnot1.loc[9,'total']
        self.df_recoSum.loc[6, 'Dineout'] = self.df_Dnot1.loc[14,'total']
        self.df_recoSum.loc[8, 'Dineout'] = self.df_Dnot1.loc[22,'total']
        self.df_recoSum.loc[10, 'Dineout'] = self.df_Dnot1.loc[24,'total']
        self.df_recoSum.loc[11, 'Dineout'] = self.df_Dnot1.loc[25,'total']
        self.df_recoSum.loc[12, 'Dineout'] = self.df_recoSum.loc[10, 'Dineout'] - self.df_recoSum.loc[11, 'Dineout']
        ### TOTAL SUMMARY
        self.df_recoSum.loc[0, 'Total'] = self.df_recoSum.loc[0, 'Swiggy'] + self.df_recoSum.loc[0, 'Zomato'] + self.df_recoSum.loc[0, 'Dunzo'] \
                                          + self.df_recoSum.loc[0, 'Nearbuy'] + self.df_recoSum.loc[0, 'Dineout']
        self.df_recoSum.loc[1, 'Total'] = self.df_recoSum.loc[1, 'Swiggy'] + self.df_recoSum.loc[1, 'Zomato'] + self.df_recoSum.loc[1, 'Dunzo'] \
                                          + self.df_recoSum.loc[1, 'Nearbuy'] + self.df_recoSum.loc[1, 'Dineout']
        self.df_recoSum.loc[2, 'Total'] = self.df_recoSum.loc[0, 'Total'] - self.df_recoSum.loc[1, 'Total']
        self.df_recoSum.loc[4, 'Total'] = self.df_recoSum.loc[4, 'Swiggy'] + self.df_recoSum.loc[4, 'Zomato'] + self.df_recoSum.loc[4, 'Dunzo'] \
                                          + self.df_recoSum.loc[4, 'Nearbuy'] + self.df_recoSum.loc[4, 'Dineout']
        self.df_recoSum.loc[6, 'Total'] = self.df_recoSum.loc[6, 'Swiggy'] + self.df_recoSum.loc[6, 'Zomato'] + self.df_recoSum.loc[6, 'Dunzo'] \
                                          + self.df_recoSum.loc[6, 'Nearbuy'] + self.df_recoSum.loc[6, 'Dineout']
        self.df_recoSum.loc[8, 'Total'] = self.df_recoSum.loc[8, 'Swiggy'] + self.df_recoSum.loc[8, 'Zomato'] + self.df_recoSum.loc[8, 'Dunzo'] \
                                          + self.df_recoSum.loc[8, 'Nearbuy'] + self.df_recoSum.loc[8, 'Dineout']
        self.df_recoSum.loc[10, 'Total'] = self.df_recoSum.loc[0, 'Total'] + self.df_recoSum.loc[4, 'Total'] + self.df_recoSum.loc[6, 'Total'] + self.df_recoSum.loc[8, 'Total']
        self.df_recoSum.loc[11, 'Total'] = self.df_recoSum.loc[11, 'Swiggy'] + self.df_recoSum.loc[11, 'Zomato'] + self.df_recoSum.loc[11, 'Dunzo'] \
                                           + self.df_recoSum.loc[11, 'Nearbuy'] + self.df_recoSum.loc[11, 'Dineout']
        self.df_recoSum.loc[12, 'Total'] = self.df_recoSum.loc[10, 'Total'] - self.df_recoSum.loc[11, 'Total']

        pList = [0,1,2,4,6,8,10,11,12]
        for i in pList:
            self.df_recoSum.loc[i,'% to total Swiggy sales'] = (self.df_recoSum.loc[i,'Swiggy']/self.df_recoSum.loc[0,'Swiggy']) * 100
            self.df_recoSum.loc[i,'% to total Zomato sales'] = (self.df_recoSum.loc[i,'Zomato']/self.df_recoSum.loc[0,'Zomato']) * 100
            self.df_recoSum.loc[i,'% to total Dunzo sales'] = (self.df_recoSum.loc[i,'Dunzo']/self.df_recoSum.loc[0,'Dunzo']) * 100
            self.df_recoSum.loc[i,'% to total Nearbuy sales'] = (self.df_recoSum.loc[i,'Nearbuy']/self.df_recoSum.loc[0,'Nearbuy']) * 100
            self.df_recoSum.loc[i,'% to total Dineout sales'] = (self.df_recoSum.loc[i,'Dineout']/self.df_recoSum.loc[0,'Dineout']) * 100
            self.df_recoSum.loc[i,'% to Total sales of all channels'] = (self.df_recoSum.loc[i,'Total']/self.df_recoSum.loc[0,'Total']) * 100


        #####------reco-summary end-------------#######
        self.table['reco_summary'] = self.df_recoSum.to_dict(orient="records")
        self.table['swiggy_reco'] = self.df_swgy1.to_dict(orient="records")
        self.table['zomato_reco'] = self.df_zom1.to_dict(orient="records")
        self.table['dunzo_reco'] = self.df_Dunz1.to_dict(orient="records")
        self.table['nearbuy_reco'] = self.df_Near1.to_dict(orient="records")
        self.table['dineout_reco'] = self.df_Dnot1.to_dict(orient="records")

        table_d1 = collections.OrderedDict(self.table)
        tableitems = list(table_d1.items())
        keys1 = ['Particulars', 'total', 'percentage']
        L1 = []
        L2 = []
        L3 = []
        L4 = []
        L5 = []
        for obj in tableitems[1][1]:
            L1.append({'fields':{x:obj[x] for x in keys1}, 'details':{x:obj[x] for x in obj if x not in keys1}})
        for j in L1:
            j['details'] = [{'key':k, 'value':v} for idx, (k,v) in enumerate(list(j['details'].items()))]
        for obj in tableitems[2][1]:
            L2.append({'fields':{x:obj[x] for x in keys1}, 'details':{x:obj[x] for x in obj if x not in keys1}})
        for j in L2:
            j['details'] = [{'key':k, 'value':v} for idx, (k,v) in enumerate(list(j['details'].items()))]
        for obj in tableitems[3][1]:
            L3.append({'fields':{x:obj[x] for x in keys1}, 'details':{x:obj[x] for x in obj if x not in keys1}})
        for j in L3:
            j['details'] = [{'key':k, 'value':v} for idx, (k,v) in enumerate(list(j['details'].items()))]
        for obj in tableitems[4][1]:
            L4.append({'fields':{x:obj[x] for x in keys1}, 'details':{x:obj[x] for x in obj if x not in keys1}})
        for j in L4:
            j['details'] = [{'key':k, 'value':v} for idx, (k,v) in enumerate(list(j['details'].items()))]
        for obj in tableitems[5][1]:
            L5.append({'fields':{x:obj[x] for x in keys1}, 'details':{x:obj[x] for x in obj if x not in keys1}})
        for j in L5:
            j['details'] = [{'key':k, 'value':v} for idx, (k,v) in enumerate(list(j['details'].items()))]



        d = datetime.now().astimezone(timezone('Asia/Kolkata'))
        mon=d.strftime("%b")
        year=d.strftime("%Y")
        mn_yr = mon + ' ' +year
        rep_name = "Sales Channel Reconciliation Report"
        head_json = {"customer":"Arbit Max Dude", "report_type":rep_name, "month_year":mn_yr }
        self.json = simplejson.dumps({"head":head_json,"reco_summary":self.table['reco_summary'], "swiggy_reco":L1, "zomato_reco":L2, "dunzo_reco":L3, "nearbuy_reco":L4, "dineout_reco":L5},ignore_nan=True)


#----------------Functions--------------------#
def merge_cells(workbook,worksheet,cell1,cell2,string=None):
    try:
        merge_center = workbook.add_format({
            'bold': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font': 13})
        worksheet.merge_range(str(cell1)+':'+str(cell2), string, merge_center)

    except Exception as e:
        print(e)

def hide_gridlines(worksheet,num):
    try:
        worksheet.hide_gridlines(num)
    except Exception as e:
        print(e)

def set_column_width(worksheet,A,B,num):
    try:
        worksheet.set_column(str(A)+':'+str(B), num)
    except Exception as e:
        print(e)

def insert_image(worksheet,image,cell):
    worksheet.insert_image(cell, image,
                           {'x_offset': 15, 'y_offset': 0,
                            'x_scale': 0.5, 'y_scale': 0.5})


def set_bold(worksheet,A):
    pass

def set_format(worksheet,A):
    pass

def draw_border(worksheet,A):
    pass
